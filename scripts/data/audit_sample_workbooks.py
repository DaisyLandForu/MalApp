from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DIRECTORIES = {
    "malicious": ROOT / "output_app_judgment",
    "white": ROOT / "output_app_white_lite",
}
BATCH_RE = re.compile(r"第(\d+)批_共(\d+)批_(\d{8}_\d{6})")


def clean(value: Any) -> str:
    return str(value or "").strip().upper()


def inspect_file(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    file_md5s: set[str] = set()
    invalid = 0
    duplicate_rows = 0
    labels = Counter()
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = [str(item or "").strip() for item in next(rows, ())]
            md5_index = next(
                (index for index, value in enumerate(headers) if value.lower() == "md5"),
                None,
            )
            big_index = next(
                (
                    index
                    for index, value in enumerate(headers)
                    if value in {"input_fraudGaType", "涉诈应用大类"}
                ),
                None,
            )
            sheet_md5s = set()
            row_count = 0
            for values in rows:
                row_count += 1
                if md5_index is None or md5_index >= len(values):
                    invalid += 1
                    continue
                md5 = clean(values[md5_index])
                if not re.fullmatch(r"[A-F0-9]{32}", md5):
                    invalid += 1
                    continue
                if md5 in sheet_md5s:
                    duplicate_rows += 1
                sheet_md5s.add(md5)
                file_md5s.add(md5)
                if big_index is not None and big_index < len(values):
                    labels[str(values[big_index] or "").strip()] += 1
            sheets.append(
                {
                    "name": sheet.title,
                    "rows": row_count,
                    "columns": len(headers),
                    "headers": headers,
                    "unique_md5": len(sheet_md5s),
                }
            )
    finally:
        workbook.close()
    batch = BATCH_RE.search(path.stem)
    return {
        "file": path.name,
        "path": str(path),
        "size_bytes": path.stat().st_size,
        "batch_number": int(batch.group(1)) if batch else None,
        "batch_total": int(batch.group(2)) if batch else None,
        "export_id": batch.group(3) if batch else "",
        "sheets": sheets,
        "unique_md5": len(file_md5s),
        "invalid_md5_rows": invalid,
        "duplicate_rows_inside_file": duplicate_rows,
        "label_counts": dict(labels),
        "_md5s": file_md5s,
    }


def inspect_directory(name: str, directory: Path) -> dict[str, Any]:
    files = [inspect_file(path) for path in sorted(directory.glob("*.xlsx"))]
    sets = {item["file"]: item["_md5s"] for item in files}
    overlaps = []
    for index, left in enumerate(files):
        for right in files[index + 1 :]:
            intersection = sets[left["file"]] & sets[right["file"]]
            if intersection:
                overlaps.append(
                    {
                        "left": left["file"],
                        "right": right["file"],
                        "overlap": len(intersection),
                        "left_coverage": round(
                            len(intersection) / max(1, left["unique_md5"]), 6
                        ),
                        "right_coverage": round(
                            len(intersection) / max(1, right["unique_md5"]), 6
                        ),
                    }
                )
    export_groups: dict[str, list[dict[str, Any]]] = {}
    for item in files:
        export_groups.setdefault(item["export_id"] or item["file"], []).append(item)
    exports = []
    for export_id, items in export_groups.items():
        union = set().union(*(item["_md5s"] for item in items))
        expected = max((item["batch_total"] or 1 for item in items), default=1)
        observed_batches = sorted(
            item["batch_number"] for item in items if item["batch_number"] is not None
        )
        exports.append(
            {
                "export_id": export_id,
                "files": [item["file"] for item in items],
                "expected_batches": expected,
                "observed_batches": observed_batches,
                "complete": observed_batches == list(range(1, expected + 1)),
                "unique_md5": len(union),
                "_md5s": union,
            }
        )
    export_overlaps = []
    for index, left in enumerate(exports):
        for right in exports[index + 1 :]:
            intersection = left["_md5s"] & right["_md5s"]
            export_overlaps.append(
                {
                    "left": left["export_id"],
                    "right": right["export_id"],
                    "overlap": len(intersection),
                    "left_coverage": round(
                        len(intersection) / max(1, left["unique_md5"]), 6
                    ),
                    "right_coverage": round(
                        len(intersection) / max(1, right["unique_md5"]), 6
                    ),
                }
            )
    recommended = max(
        (item for item in exports if item["complete"]),
        key=lambda item: item["unique_md5"],
        default=max(exports, key=lambda item: item["unique_md5"], default=None),
    )
    for item in files:
        item.pop("_md5s", None)
    for item in exports:
        item.pop("_md5s", None)
    return {
        "name": name,
        "directory": str(directory),
        "files": files,
        "file_overlaps": overlaps,
        "exports": exports,
        "export_overlaps": export_overlaps,
        "recommended_export_id": recommended["export_id"] if recommended else "",
        "recommended_files": recommended["files"] if recommended else [],
        "recommended_unique_md5": recommended["unique_md5"] if recommended else 0,
    }


def main() -> None:
    report = {
        name: inspect_directory(name, directory)
        for name, directory in DIRECTORIES.items()
    }
    malicious = set()
    white = set()
    for name, target in (("malicious", malicious), ("white", white)):
        selected = set(report[name]["recommended_files"])
        for path in DIRECTORIES[name].glob("*.xlsx"):
            if path.name not in selected:
                continue
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    rows = sheet.iter_rows(values_only=True)
                    headers = [str(item or "").strip() for item in next(rows, ())]
                    index = next(
                        (
                            idx
                            for idx, value in enumerate(headers)
                            if value.lower() == "md5"
                        ),
                        None,
                    )
                    if index is None:
                        continue
                    for values in rows:
                        if index < len(values):
                            md5 = clean(values[index])
                            if re.fullmatch(r"[A-F0-9]{32}", md5):
                                target.add(md5)
            finally:
                workbook.close()
    report["cross_label_overlap"] = {
        "count": len(malicious & white),
        "examples": sorted(malicious & white)[:30],
        "malicious_unique": len(malicious),
        "white_unique": len(white),
    }
    output = ROOT / "training_artifacts" / "sample_workbook_audit.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
