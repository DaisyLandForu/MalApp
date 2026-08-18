from __future__ import annotations

import argparse
import csv
import hashlib
import json
import shutil
import sqlite3
import sys
import time
from pathlib import Path
from typing import Any

import numpy as np
import xgboost as xgb
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from malapp.governance.artifacts import build_xgboost_manifest  # noqa: E402
from malapp.governance.leakage import require_training_clearance  # noqa: E402
from training.xgboost import pipeline as xp  # noqa: E402

DEFAULT_OUT = ROOT / "training_artifacts" / "xgb_selected_20260616"
APP_SHEET = "APP\u7814\u5224\u6570\u636e"
AGENTS = xp.AGENTS
FEATURES = xp.FEATURES
APP_EXTRA_HEADERS = [
    "xgb_probability",
    "xgb_verdict",
    "split",
    "label_source",
    "sample_weight",
]


def stable_order_key(md5: str, salt: str) -> str:
    return hashlib.sha1(f"{salt}:{md5}".encode("utf-8")).hexdigest()


def split_train_stack_val(md5: str) -> str:
    bucket = int(hashlib.sha1(md5.encode("ascii", errors="ignore")).hexdigest()[:8], 16) % 100
    if bucket < 70:
        return "train"
    if bucket < 85:
        return "stack"
    return "val"


def clean_text(value: Any) -> str:
    if value in (None, "", "None", "nan", "NaN"):
        return ""
    return str(value).strip()


def truthy_label(value: Any) -> int:
    text = clean_text(value).lower()
    return 1 if text in {"true", "1", "yes", "malicious", "\u6076\u610f"} else 0


def engine_label_from_score(score: Any) -> str:
    try:
        number = float(str(score).strip())
    except Exception:
        return ""
    if number >= 90:
        return "malicious"
    if number > 0:
        return "suspicious"
    return "benign"


def read_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[APP_SHEET] if APP_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = [str(item or "").strip() for item in next(rows)]
        result = []
        for row in rows:
            item = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
            md5 = clean_text(item.get("md5")).upper()
            if md5:
                item["md5"] = md5
                result.append(item)
        return result
    finally:
        wb.close()


def app_row_to_payload(row: dict[str, Any]) -> dict[str, Any]:
    payload = {
        "md5": row.get("md5"),
        "sample_id": row.get("sample_id") or row.get("md5"),
        "app_name": row.get("app_name"),
        "sample_app_id": row.get("package_name"),
        "sha1": row.get("sha1"),
        "sha256": row.get("sha256"),
        "file_size": row.get("file_size"),
        "file_type": row.get("file_type"),
        "platform": row.get("file_type") or "Android",
        "certificate_fingerprint": row.get("certificate_fingerprint"),
        "cert_sha1": row.get("cert_sha1"),
        "cert_sha256": row.get("cert_sha256"),
        "certificate_owner": row.get("certificate_owner"),
        "certificate_developer": row.get("certificate_developer"),
        "sdk_list": row.get("sdk_list") or row.get("plugins"),
        "packer": row.get("packer") or row.get("code_fuscator"),
        "fake_app": row.get("fake_app"),
        "impersonation_flag": row.get("fake_app"),
        "app_version_status": row.get("rebuild_type"),
        "official_app_name": row.get("official_app_name"),
        "virus_name": row.get("virus_name"),
        "virus_description": row.get("virus_description"),
        "harm_type": row.get("risk_type"),
        "anti_fraud_tag": row.get("fraud_type_info"),
        "fraud_flag": row.get("fraud_flag") or row.get("source_malicious"),
        "fraud_category_big": row.get("fraud_category_big"),
        "fraud_category_small": row.get("fraud_category_small"),
        "fraud_family": row.get("fraud_name"),
        "control_url": row.get("control_url"),
        "download_url": row.get("download_url"),
        "domains": row.get("domains"),
        "ips": row.get("ips"),
        "label_source": row.get("label_source"),
    }
    if clean_text(row.get("engine_360_score")):
        payload["engine_a_score"] = row.get("engine_360_score")
        payload["engine_a_label"] = engine_label_from_score(row.get("engine_360_score"))
    if clean_text(row.get("engine_cm_score")):
        payload["engine_b_score"] = row.get("engine_cm_score")
        payload["engine_b_label"] = engine_label_from_score(row.get("engine_cm_score"))
    return {key: value for key, value in payload.items() if value not in ("", None)}


def load_reference_sets() -> dict[str, set[str]]:
    db_path = ROOT / "training_artifacts" / "xgb" / "xgb_training.db"
    if not db_path.exists():
        return {key: set() for key in ("md5", "name", "package", "sign_md5", "sign_sha1", "sign_sha256")}
    conn = sqlite3.connect(db_path)
    try:
        return xp.load_reference_sets(conn)
    finally:
        conn.close()


def build_dataset(paths: dict[str, Path]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    specs = [
        ("known_malicious", paths["malicious"], 1, 1.0, None, "base_malicious"),
        ("known_white", paths["white"], 0, 1.0, None, "base_white"),
        ("manual_conflict", paths["manual"], 1, 0.9, 3000, "manual_conflict"),
        ("consensus_score0", paths["consensus"], 0, 0.9, 800, "consensus_score0"),
    ]
    train_like: list[dict[str, Any]] = []
    heldout: list[dict[str, Any]] = []
    for source, path, label, weight, train_limit, salt in specs:
        rows = read_workbook(path)
        rows.sort(key=lambda item: stable_order_key(str(item["md5"]), salt))
        for index, row in enumerate(rows):
            if train_limit is not None and index >= train_limit:
                target_split = "test"
                target = heldout
            else:
                target_split = split_train_stack_val(row["md5"])
                target = train_like
            payload = app_row_to_payload(row)
            payload["md5"] = row["md5"]
            if source in {"manual_conflict", "consensus_score0"}:
                label = truthy_label(row.get("source_malicious"))
            target.append(
                {
                    "md5": row["md5"],
                    "label": int(label),
                    "label_source": source,
                    "sample_weight": float(weight),
                    "split": target_split,
                    "payload": payload,
                    "raw": row,
                }
            )
    return train_like, heldout


def init_db(path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS xgb_samples (
            md5 TEXT PRIMARY KEY, label INTEGER NOT NULL, label_source TEXT NOT NULL,
            sample_weight REAL NOT NULL, split TEXT NOT NULL, group_id TEXT NOT NULL,
            payload_json TEXT NOT NULL, features_json TEXT NOT NULL
        );
        """
    )
    conn.execute("DELETE FROM xgb_samples")
    return conn


def materialize(out_dir: Path, records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    refs = load_reference_sets()
    db_path = out_dir / "xgb_training.db"
    conn = init_db(db_path)
    rows_for_train: list[dict[str, Any]] = []
    stats: dict[str, Any] = {"splits": {}, "label_sources": {}, "labels": {}}
    try:
        batch = []
        for record in records:
            md5 = record["md5"]
            features = xp.feature_vector(md5, record["payload"], refs)
            group_id = md5
            batch.append(
                (
                    md5,
                    record["label"],
                    record["label_source"],
                    record["sample_weight"],
                    record["split"],
                    group_id,
                    json.dumps(record["payload"], ensure_ascii=False),
                    json.dumps(features, ensure_ascii=False),
                )
            )
            rows_for_train.append(
                {
                    "md5": md5,
                    "label": record["label"],
                    "label_source": record["label_source"],
                    "sample_weight": record["sample_weight"],
                    "split": record["split"],
                    **features,
                }
            )
            stats["splits"][record["split"]] = stats["splits"].get(record["split"], 0) + 1
            stats["label_sources"][record["label_source"]] = stats["label_sources"].get(record["label_source"], 0) + 1
            stats["labels"][str(record["label"])] = stats["labels"].get(str(record["label"]), 0) + 1
        conn.executemany("INSERT OR REPLACE INTO xgb_samples VALUES (?,?,?,?,?,?,?,?)", batch)
        conn.commit()
    finally:
        conn.close()
    shutil.copyfile(db_path, out_dir / "runtime_training.db")
    return rows_for_train, stats


def arrays(rows: list[dict[str, Any]], features: list[str]):
    x_values = np.asarray([[float(row.get(feature, 0.0)) for feature in features] for row in rows], dtype=np.float32)
    labels = np.asarray([row["label"] for row in rows], dtype=np.int32)
    weights = np.asarray([row.get("sample_weight", 1.0) for row in rows], dtype=np.float32)
    return x_values, labels, weights


def fit_model(rows: list[dict[str, Any]], features: list[str], name: str, model_dir: Path, curve_dir: Path, train_split: str = "train"):
    train_rows = [row for row in rows if row["split"] == train_split]
    val_rows = [row for row in rows if row["split"] == "val"]
    x_train, y_train, w_train = arrays(train_rows, features)
    x_val, y_val, _ = arrays(val_rows, features)
    model = xgb.XGBClassifier(
        n_estimators=900,
        max_depth=4,
        learning_rate=0.04,
        min_child_weight=4,
        subsample=0.85,
        colsample_bytree=0.9,
        reg_alpha=0.1,
        reg_lambda=1.5,
        objective="binary:logistic",
        eval_metric="logloss",
        tree_method="hist",
        random_state=20260616,
        n_jobs=8,
        early_stopping_rounds=50,
    )
    start = time.perf_counter()
    model.fit(x_train, y_train, sample_weight=w_train, eval_set=[(x_train, y_train), (x_val, y_val)], verbose=False)
    seconds = round(time.perf_counter() - start, 3)
    history = model.evals_result()
    model.save_model(model_dir / f"{name}.json")
    xp.save_curve(curve_dir / f"{name}_loss.png", history, f"{name} logloss")
    metrics = {
        "duration_seconds": seconds,
        "train": xp.metrics_for(model, train_rows, features),
        "validation": xp.metrics_for(model, val_rows, features),
        "test": xp.metrics_for(model, [row for row in rows if row["split"] == "test"], features),
        "best_iteration": int(model.best_iteration),
        "final_train_logloss": history["validation_0"]["logloss"][model.best_iteration],
        "final_validation_logloss": history["validation_1"]["logloss"][model.best_iteration],
        "feature_importance": sorted(zip(features, model.feature_importances_.tolist(), strict=False), key=lambda item: item[1], reverse=True),
        "params": model.get_params(),
    }
    print(f"[selected-xgb] {name}: {seconds}s best={model.best_iteration} val_logloss={metrics['final_validation_logloss']:.6f}", flush=True)
    return model, metrics


def stack_records(rows: list[dict[str, Any]], models: dict[str, Any]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        current = dict(row)
        for agent, model in models.items():
            values = np.asarray([[float(row.get(feature, 0.0)) for feature in FEATURES[agent]]], dtype=np.float32)
            current[f"{agent}_prob"] = float(model.predict_proba(values)[:, 1][0])
        result.append(current)
    return result


def add_wec_features(rows: list[dict[str, Any]], fusion, fusion_features: list[str]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        current = dict(row)
        values = np.asarray([[float(row.get(feature, 0.0)) for feature in fusion_features]], dtype=np.float32)
        current["engine_c_prob"] = float(fusion.predict_proba(values)[:, 1][0])
        result.append(current)
    return result


def write_manifest(
    out_dir: Path,
    thresholds: dict[str, Any],
    metrics: dict[str, Any],
) -> None:
    manifest = build_xgboost_manifest(
        model_dir=out_dir / "models",
        database_path=out_dir / "xgb_training.db",
        agents={agent: FEATURES[agent] for agent in AGENTS},
        fusion_features=[f"{agent}_prob" for agent in AGENTS],
        wec_features=[
            "engine_a_prob",
            "engine_b_prob",
            "engine_c_prob",
            "engine_score_gap",
            "engine_disagreement",
        ],
        thresholds=thresholds,
        metrics=metrics,
        model_version="xgb-selected-v1",
    )
    (out_dir / "models" / "runtime_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def export_rows(out_dir: Path, heldout: list[dict[str, Any]], wec_rows: list[dict[str, Any]], wec, wec_features: list[str], thresholds: dict[str, Any]) -> None:
    output_rows = []
    test_rows = [row for row in wec_rows if row["split"] == "test"]
    x_test, _, _ = arrays(test_rows, wec_features)
    probabilities = wec.predict_proba(x_test)[:, 1] if len(test_rows) else []
    prob_by_md5 = {row["md5"]: float(prob) for row, prob in zip(test_rows, probabilities, strict=False)}
    for item in heldout:
        raw = dict(item["raw"])
        prob = prob_by_md5.get(item["md5"], 0.5)
        verdict = "benign" if prob < thresholds["benign_threshold"] else "malicious" if prob >= thresholds["malicious_threshold"] else "suspicious"
        raw.update(
            {
                "xgb_probability": round(prob, 6),
                "xgb_verdict": verdict,
                "split": "test",
                "label_source": item["label_source"],
                "sample_weight": item["sample_weight"],
            }
        )
        output_rows.append(raw)
    if not output_rows:
        return
    headers = list(output_rows[0].keys())
    for header in APP_EXTRA_HEADERS:
        if header not in headers:
            headers.append(header)
    csv_path = out_dir / "test_set_for_app.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(output_rows)
    wb = Workbook()
    ws = wb.active
    ws.title = APP_SHEET
    ws.append(headers)
    for row in output_rows:
        ws.append([row.get(header, "") for header in headers])
    wb.save(out_dir / "test_set_for_app.xlsx")


def train(paths: dict[str, Path], out_dir: Path) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "models").mkdir(exist_ok=True)
    (out_dir / "curves").mkdir(exist_ok=True)
    total_start = time.perf_counter()
    train_like, heldout = build_dataset(paths)
    all_records = train_like + heldout
    rows, prep_stats = materialize(out_dir, all_records)
    report: dict[str, Any] = {
        "data_plan": {
            "base_malicious": 5000,
            "base_white": 2000,
            "manual_conflict_train_or_validation": 3000,
            "manual_conflict_test": len([x for x in heldout if x["label_source"] == "manual_conflict"]),
            "consensus_score0_train_or_validation": 800,
            "consensus_score0_test": len([x for x in heldout if x["label_source"] == "consensus_score0"]),
        },
        "prepare": prep_stats,
        "models": {},
    }
    models = {}
    for agent in AGENTS:
        model, metrics = fit_model(rows, FEATURES[agent], agent, out_dir / "models", out_dir / "curves", "train")
        models[agent] = model
        report["models"][agent] = metrics
    stacked = stack_records(rows, models)
    fusion_features = [f"{agent}_prob" for agent in AGENTS]
    fusion, fusion_metrics = fit_model(stacked, fusion_features, "fusion", out_dir / "models", out_dir / "curves", "stack")
    report["models"]["fusion"] = fusion_metrics
    wec_rows = add_wec_features(stacked, fusion, fusion_features)
    wec_features = ["engine_a_prob", "engine_b_prob", "engine_c_prob", "engine_score_gap", "engine_disagreement"]
    wec, wec_metrics = fit_model(wec_rows, wec_features, "wec", out_dir / "models", out_dir / "curves", "stack")
    report["models"]["wec"] = wec_metrics
    thresholds = xp.learn_thresholds(wec_rows, wec, wec_features)
    report["thresholds"] = thresholds
    report["test"] = xp.evaluate_split(wec_rows, wec, wec_features, "test", thresholds)
    (out_dir / "models" / "thresholds.json").write_text(json.dumps(thresholds, ensure_ascii=False, indent=2), encoding="utf-8")
    write_manifest(out_dir, thresholds, report)
    export_rows(out_dir, heldout, wec_rows, wec, wec_features, thresholds)
    report["duration_seconds_total"] = round(time.perf_counter() - total_start, 3)
    (out_dir / "training_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--malicious", required=True)
    parser.add_argument("--white", required=True)
    parser.add_argument("--manual", required=True)
    parser.add_argument("--consensus", required=True)
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    parser.add_argument("--dataset-manifest", required=True)
    args = parser.parse_args()
    require_training_clearance(Path(args.dataset_manifest))
    paths = {
        "malicious": Path(args.malicious),
        "white": Path(args.white),
        "manual": Path(args.manual),
        "consensus": Path(args.consensus),
    }
    report = train(paths, Path(args.out))
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
