from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_HEADERS = [
    "md5",
    "sample_id",
    "app_name",
    "package_name",
    "sha1",
    "sha256",
    "version_name",
    "version_code",
    "file_size",
    "file_type",
    "signature_status",
    "certificate_fingerprint",
    "cert_sha1",
    "cert_sha256",
    "certificate_owner",
    "certificate_developer",
    "permissions",
    "plugins",
    "sdk_list",
    "packer",
    "code_fuscator",
    "unshell_info",
    "fake_app",
    "genuine",
    "official_app_name",
    "rebuild_type",
    "icon_md5",
    "virus_name",
    "risk_type",
    "virus_description",
    "fraud_flag",
    "fraud_type_info",
    "fraud_name",
    "fraud_category",
    "fraud_category_big",
    "fraud_category_small",
    "control_url",
    "download_url",
    "lt_urls",
    "dynamic_nets",
    "domains",
    "top_domains",
    "ips",
    "countries",
    "operators",
    "domain_count",
    "ip_count",
    "network_record_count",
    "source_risk_score",
    "source_malicious",
    "source_violation",
    "query_time",
]


FIELD_NOTES = {
    "md5": ("原始输入/接口1.md5", "样本唯一标识和历史报告复用"),
    "app_name": ("接口1.appName/input_appName", "应用名称和业务语义分析"),
    "package_name": ("接口1.pkgName", "包名异常、变种和仿冒分析"),
    "sha1": ("接口1.fileSha1", "样本身份信息"),
    "sha256": ("接口1.fileSha256", "样本身份信息"),
    "signature_status": ("接口1.signature", "是否存在签名；不伪造签名合法性结论"),
    "certificate_fingerprint": ("接口1.signature.md5", "签名和正版库匹配"),
    "permissions": ("接口1.permissions", "静态智能体高危权限分析"),
    "sdk_list": ("接口1.plugins.plugName", "第三方 SDK 风险分析"),
    "packer": ("保守转换", "没有可靠壳结论时设为否，避免把混淆分误判为加固"),
    "code_fuscator": ("接口1.codeFuscator", "保留原始代码混淆指标"),
    "fake_app": ("接口1.fakeApp", "仿冒智能体输入"),
    "genuine": ("接口1.genuine", "正版应用参照"),
    "virus_name": ("接口1.codeSafeResultVirusName", "恶意家族与业务危害标签"),
    "risk_type": ("接口1.codeSafeVirusType", "危害类型"),
    "fraud_type_info": ("接口1.fraudTypeInfo", "业务标签与危害链分析"),
    "fraud_name": ("接口1.fraudName", "欺诈家族"),
    "fraud_category_big": ("input_fraudGaType", "业务大类"),
    "fraud_category_small": ("input_fraudGaSubType", "业务小类"),
    "lt_urls": ("接口1.ltUrls", "提取并压缩后的通联 URL"),
    "dynamic_nets": ("接口1.dynamicNets", "提取并压缩后的动态通联 URL"),
    "domains": ("MD5通联汇总.domains", "情报溯源智能体域名输入"),
    "top_domains": ("MD5通联汇总.top_domains", "多层域名分析"),
    "ips": ("MD5通联汇总.ips", "情报溯源智能体 IP 输入"),
    "countries": ("MD5通联汇总.countries", "IP 归属信息"),
    "operators": ("MD5通联汇总.operators", "运营商与云服务商信息"),
    "source_risk_score": ("接口1.score", "仅保留原系统分数，不映射为 APP 最终结论"),
    "source_malicious": ("接口1.malicious", "仅保留来源系统标记，不作为人工金标"),
}


def clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\x00", "").strip()
    return value


def parse_json(value: Any, default: Any) -> Any:
    if isinstance(value, (dict, list)):
        return value
    text = str(value or "").strip()
    if not text:
        return default
    try:
        return json.loads(text)
    except (json.JSONDecodeError, TypeError):
        return default


def json_list(items: list[Any]) -> str:
    result: list[str] = []
    seen: set[str] = set()
    for item in items:
        text = str(item or "").strip()
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return json.dumps(result, ensure_ascii=False, separators=(",", ":"))


def permission_names(value: Any) -> str:
    rows = parse_json(value, [])
    if not isinstance(rows, list):
        return "[]"
    return json_list(
        row.get("name") if isinstance(row, dict) else row
        for row in rows
    )


def plugin_names(value: Any) -> tuple[str, str]:
    rows = parse_json(value, [])
    if not isinstance(rows, list):
        return "[]", "[]"
    names = [
        row.get("plugName") if isinstance(row, dict) else row
        for row in rows
    ]
    compact = json_list(names)
    return compact, compact


def communication_urls(value: Any, *, dynamic: bool = False) -> str:
    rows = parse_json(value, [])
    if not isinstance(rows, list):
        return "[]"
    keys = ("url", "host", "domain", "ip") if dynamic else ("url", "domain", "ip")
    items: list[str] = []
    for row in rows:
        if not isinstance(row, dict):
            items.append(str(row))
            continue
        for key in keys:
            if row.get(key):
                items.append(str(row[key]))
                break
    return json_list(items[:300])


def bool_text(value: Any) -> str:
    return "是" if str(value or "").strip().lower() in {"1", "true", "yes", "y", "是"} else "否"


def signature_fields(value: Any) -> dict[str, str]:
    signature = parse_json(value, {})
    if not isinstance(signature, dict) or not signature:
        return {
            "signature_status": "missing",
            "certificate_fingerprint": "",
            "cert_sha1": "",
            "cert_sha256": "",
            "certificate_owner": "",
            "certificate_developer": "",
        }
    return {
        "signature_status": "present",
        "certificate_fingerprint": str(signature.get("md5") or ""),
        "cert_sha1": str(signature.get("sha1") or ""),
        "cert_sha256": str(signature.get("sha256") or ""),
        "certificate_owner": str(signature.get("owner") or ""),
        "certificate_developer": str(signature.get("developer") or ""),
    }


def indexed_rows(sheet) -> dict[str, dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    result: dict[str, dict[str, Any]] = {}
    for values in rows:
        record = {headers[index]: clean(value) for index, value in enumerate(values)}
        md5 = str(record.get("md5") or "").upper()
        if md5:
            result[md5] = record
    return result


def convert(input_path: Path, output_path: Path) -> dict[str, int]:
    source = load_workbook(input_path, read_only=True, data_only=True)
    try:
        main_rows = indexed_rows(source.worksheets[1])
        network_rows = indexed_rows(source.worksheets[2])
    finally:
        source.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "APP研判数据"
    sheet.append(OUTPUT_HEADERS)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for md5, raw in main_rows.items():
        network = network_rows.get(md5, {})
        signature = signature_fields(raw.get("signature"))
        plugins, sdk_list = plugin_names(raw.get("plugins"))
        record = {
            "md5": md5,
            "sample_id": md5,
            "app_name": raw.get("appName") or raw.get("input_appName"),
            "package_name": raw.get("pkgName"),
            "sha1": raw.get("fileSha1"),
            "sha256": raw.get("fileSha256"),
            "version_name": raw.get("versionName") or raw.get("version"),
            "version_code": raw.get("versionCode"),
            "file_size": raw.get("size"),
            "file_type": raw.get("fileType"),
            **signature,
            "permissions": permission_names(raw.get("permissions")),
            "plugins": plugins,
            "sdk_list": sdk_list,
            "packer": "否",
            "code_fuscator": raw.get("codeFuscator"),
            "unshell_info": str(raw.get("unshellInfo") or "")[:32000],
            "fake_app": bool_text(raw.get("fakeApp")),
            "genuine": raw.get("genuine"),
            "official_app_name": raw.get("genuine"),
            "rebuild_type": raw.get("rebuildType"),
            "icon_md5": raw.get("iconMd5") or raw.get("iconFingerMd5"),
            "virus_name": raw.get("codeSafeResultVirusName"),
            "risk_type": raw.get("codeSafeVirusType"),
            "virus_description": raw.get("riskEvaluateResult"),
            "fraud_flag": bool_text(raw.get("maliciousOrViolation")),
            "fraud_type_info": str(raw.get("fraudTypeInfo") or "")[:32000],
            "fraud_name": raw.get("fraudName"),
            "fraud_category": raw.get("fraudCategory"),
            "fraud_category_big": raw.get("input_fraudGaType"),
            "fraud_category_small": raw.get("input_fraudGaSubType"),
            "control_url": raw.get("masterControlUrl"),
            "download_url": raw.get("firstShopDownUrl"),
            "lt_urls": communication_urls(raw.get("ltUrls")),
            "dynamic_nets": communication_urls(raw.get("dynamicNets"), dynamic=True),
            "domains": network.get("domains"),
            "top_domains": network.get("top_domains"),
            "ips": network.get("ips"),
            "countries": network.get("countries"),
            "operators": network.get("operators"),
            "domain_count": network.get("domain_count"),
            "ip_count": network.get("ip_count"),
            "network_record_count": network.get("url_record_count"),
            "source_risk_score": raw.get("score"),
            "source_malicious": raw.get("malicious"),
            "source_violation": raw.get("violation"),
            "query_time": raw.get("query_time"),
        }
        sheet.append([record.get(header, "") for header in OUTPUT_HEADERS])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 34,
        "B": 34,
        "C": 24,
        "D": 38,
    }
    for column in range(1, len(OUTPUT_HEADERS) + 1):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = widths.get(letter, 20)

    notes = workbook.create_sheet("字段说明")
    notes.append(["标准字段", "原始来源", "研判用途"])
    for cell in notes[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for field in OUTPUT_HEADERS:
        source_name, purpose = FIELD_NOTES.get(
            field,
            ("接口1或MD5通联汇总", "辅助信息或来源数据留存"),
        )
        notes.append([field, source_name, purpose])
    notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 32
    notes.column_dimensions["B"].width = 38
    notes.column_dimensions["C"].width = 62

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "samples": len(main_rows),
        "network_matched": sum(1 for md5 in main_rows if md5 in network_rows),
        "columns": len(OUTPUT_HEADERS),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    result = convert(args.input.resolve(), args.output.resolve())
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
