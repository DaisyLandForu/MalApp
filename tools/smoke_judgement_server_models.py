import json
import math
import os
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APP_URL = os.getenv("MALAPP_SMOKE_APP_URL", "http://127.0.0.1:8841")


def clean(value: Any) -> Any:
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def post_json(path: str, payload: dict[str, Any], timeout: int = 60) -> tuple[int, str]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        f"{APP_URL}{path}",
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.status, resp.read().decode("utf-8", "replace")
    except urllib.error.HTTPError as exc:
        return exc.code, exc.read().decode("utf-8", "replace")


def first_excel_row(path: str) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    for values in rows:
        row = {headers[index]: clean(value) for index, value in enumerate(values) if headers[index]}
        if any(value not in ("", None) for value in row.values()):
            return row
    raise RuntimeError(f"no data row found in {path}")


def main() -> None:
    settings = {
        "server_models_enabled": True,
        "model_a_api_url": "http://10.0.11.55:10000/v1",
        "model_a_api_key": "EMPTY",
        "model_a_model": "Qwen3.6-35B-A3B-FP8",
        "model_b_api_url": "http://10.0.11.82:18012/v1",
        "model_b_api_key": "EMPTY",
        "model_b_model": "malapp-model-b",
        "local_qwen_enabled": False,
    }
    status, text = post_json("/api/model/settings", settings, timeout=40)
    print("settings:", status, text[:800])

    row = first_excel_row(str(Path("training_artifacts") / "xgb_selected_20260616" / "test_set_for_app.xlsx"))
    status, text = post_json("/api/judgements", row, timeout=480)
    print("judge:", status)
    if 200 <= status < 300:
        data = json.loads(text)
        debate = data.get("debate", {})
        print("execution_mode:", debate.get("execution_mode"))
        print("debate_rounds:", debate.get("debate_rounds"))
        print("phases:", [stage.get("phase") for stage in debate.get("stages", [])])
        print("model_a_backend:", debate.get("providers", {}).get("model_a", {}).get("backend"))
        print("model_b_backend:", debate.get("providers", {}).get("model_b", {}).get("backend"))
    print(text[:5000])


if __name__ == "__main__":
    main()
