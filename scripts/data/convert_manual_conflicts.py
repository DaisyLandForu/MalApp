from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

OUTPUT_SHEET = "APP\u7814\u5224\u6570\u636e"

APP_HEADERS = [
    "md5",
    "sample_id",
    "app_name",
    "package_name",
    "sha1",
    "sha256",
    "version_name",
    "version_code",
    "file_size",
    "file_type",
    "signature_status",
    "certificate_fingerprint",
    "cert_sha1",
    "cert_sha256",
    "certificate_owner",
    "certificate_developer",
    "permissions",
    "plugins",
    "sdk_list",
    "packer",
    "code_fuscator",
    "unshell_info",
    "fake_app",
    "genuine",
    "official_app_name",
    "rebuild_type",
    "icon_md5",
    "virus_name",
    "risk_type",
    "virus_description",
    "fraud_flag",
    "fraud_type_info",
    "fraud_name",
    "fraud_category",
    "fraud_category_big",
    "fraud_category_small",
    "control_url",
    "download_url",
    "lt_urls",
    "dynamic_nets",
    "domains",
    "top_domains",
    "ips",
    "countries",
    "operators",
    "domain_count",
    "ip_count",
    "network_record_count",
    "source_risk_score",
    "source_malicious",
    "source_violation",
    "query_time",
]

EXTRA_HEADERS = [
    "gold_label",
    "label_source",
    "manual_review_result",
    "conflict_type",
    "engine_360_type",
    "engine_360_score",
    "engine_360_virus_name",
    "engine_cm_type",
    "engine_cm_score",
    "engine_cm_virus_name",
]


def clean_md5(value: Any) -> str:
    text = str(value or "").strip().upper()
    return text if re.fullmatch(r"[A-F0-9]{32}", text) else text


def clean_text(value: Any, max_chars: int = 12000) -> str:
    if value in (None, "", "None", "nan", "NaN"):
        return ""
    text = " ".join(str(value).split())
    return text[:max_chars]


def first_value(*values: Any) -> str:
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return ""


def normalize_bool(value: Any) -> str:
    text = clean_text(value).lower()
    if text in {"1", "true", "yes", "y", "\u662f"}:
        return "true"
    if text in {"0", "false", "no", "n", "\u5426"}:
        return "false"
    return ""


def row_dict(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}


def load_manual_labels(files: list[Path]) -> dict[str, dict[str, Any]]:
    labels: dict[str, dict[str, Any]] = {}
    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in wb.sheetnames:
                if "\u6c47\u603b" in sheet:
                    continue
                ws = wb[sheet]
                iterator = ws.iter_rows(values_only=True)
                try:
                    headers = [str(item or "").strip() for item in next(iterator)]
                except StopIteration:
                    continue
                for row in iterator:
                    item = row_dict(headers, row)
                    md5 = clean_md5(item.get("MD5"))
                    manual = clean_text(item.get("\u4eba\u5de5\u5ba1\u6838\u7ed3\u679c"))
                    if not md5 or not manual:
                        continue
                    labels[md5] = {
                        "md5": md5,
                        "manual_review_result": manual,
                        "conflict_type": clean_text(item.get("\u51b2\u7a81\u7c7b\u578b")),
                        "engine_360_type": clean_text(item.get("\u7c7b\u578b_360")),
                        "engine_360_score": clean_text(item.get("score_360")),
                        "engine_360_virus_name": clean_text(item.get("\u75c5\u6bd2\u540d\u79f0_360")),
                        "engine_cm_type": clean_text(item.get("\u7c7b\u578b_cm")),
                        "engine_cm_score": clean_text(item.get("score_cm")),
                        "engine_cm_virus_name": clean_text(item.get("\u75c5\u6bd2\u540d\u79f0_cm")),
                        "app_name_360": clean_text(item.get("\u5e94\u7528\u540d\u79f0_360")),
                        "app_name_cm": clean_text(item.get("\u5e94\u7528\u540d\u79f0_cm")),
                        "label_source": path.name,
                    }
        finally:
            wb.close()
    return labels


def load_engine_rows(path: Path, wanted_md5: set[str]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        iterator = ws.iter_rows(values_only=True)
        headers = [str(item or "").strip() for item in next(iterator)]
        for row in iterator:
            item = row_dict(headers, row)
            md5 = clean_md5(item.get("MD5"))
            if md5 in wanted_md5:
                result[md5] = item
    finally:
        wb.close()
    return result


def score_number(*values: Any) -> str:
    best: float | None = None
    for value in values:
        try:
            number = float(str(value).strip())
        except Exception:
            continue
        best = number if best is None else max(best, number)
    if best is None:
        return ""
    return str(int(best)) if best.is_integer() else str(best)


def manual_to_label(value: str) -> tuple[str, str]:
    text = clean_text(value)
    try:
        number = float(text)
    except Exception:
        number = 0.0
    if number >= 50:
        return "malicious", "true"
    return "benign", "false"


def convert_record(manual: dict[str, Any], row360: dict[str, Any], rowcm: dict[str, Any]) -> dict[str, Any]:
    md5 = manual["md5"]
    preferred = rowcm if rowcm else row360
    backup = row360 if rowcm else rowcm
    gold_label, source_malicious = manual_to_label(manual["manual_review_result"])
    engine_summary = {
        "360": {
            "type": manual.get("engine_360_type"),
            "score": manual.get("engine_360_score"),
            "virus_name": manual.get("engine_360_virus_name"),
            "description": clean_text(row360.get("description") or row360.get("\u75c5\u6bd2\u63cf\u8ff0")),
        },
        "cm": {
            "type": manual.get("engine_cm_type"),
            "score": manual.get("engine_cm_score"),
            "virus_name": manual.get("engine_cm_virus_name"),
            "description": clean_text(rowcm.get("description") or rowcm.get("\u75c5\u6bd2\u63cf\u8ff0")),
        },
        "manual": {
            "result": manual.get("manual_review_result"),
            "conflict_type": manual.get("conflict_type"),
        },
    }
    return {
        "md5": md5,
        "sample_id": md5,
        "app_name": first_value(preferred.get("\u5e94\u7528\u540d\u79f0"), manual.get("app_name_cm"), manual.get("app_name_360"), backup.get("\u5e94\u7528\u540d\u79f0")),
        "package_name": first_value(preferred.get("pngName"), backup.get("pngName")),
        "sha1": first_value(preferred.get("\u6837\u672csha1"), backup.get("\u6837\u672csha1")),
        "sha256": first_value(preferred.get("\u6837\u672csha256"), backup.get("\u6837\u672csha256")),
        "version_name": first_value(preferred.get("\u5e94\u7528\u7248\u672c\u53f7"), backup.get("\u5e94\u7528\u7248\u672c\u53f7")),
        "version_code": "",
        "file_size": first_value(preferred.get("size"), backup.get("size")),
        "file_type": first_value(preferred.get("\u5b89\u88c5\u5305\u6587\u4ef6\u7c7b\u578b"), preferred.get("platform"), backup.get("\u5b89\u88c5\u5305\u6587\u4ef6\u7c7b\u578b"), backup.get("platform")),
        "signature_status": "present" if first_value(preferred.get("\u8bc1\u4e66\u6307\u7eb9MD5"), backup.get("\u8bc1\u4e66\u6307\u7eb9MD5")) else "",
        "certificate_fingerprint": first_value(preferred.get("\u8bc1\u4e66\u6307\u7eb9MD5"), backup.get("\u8bc1\u4e66\u6307\u7eb9MD5")),
        "cert_sha1": first_value(preferred.get("\u5e94\u7528\u7b7e\u540d\u8bc1\u4e66sha1"), backup.get("\u5e94\u7528\u7b7e\u540d\u8bc1\u4e66sha1")),
        "cert_sha256": first_value(preferred.get("\u5e94\u7528\u7b7e\u540d\u8bc1\u4e66sha256"), backup.get("\u5e94\u7528\u7b7e\u540d\u8bc1\u4e66sha256")),
        "certificate_owner": first_value(preferred.get("\u8bc1\u4e66\u62e5\u6709\u8005"), backup.get("\u8bc1\u4e66\u62e5\u6709\u8005")),
        "certificate_developer": first_value(preferred.get("\u8bc1\u4e66\u5f00\u53d1\u8005"), preferred.get("\u5f00\u53d1\u8005\u540d\u79f0"), backup.get("\u8bc1\u4e66\u5f00\u53d1\u8005"), backup.get("\u5f00\u53d1\u8005\u540d\u79f0")),
        "permissions": "",
        "plugins": first_value(preferred.get("\u6837\u672cSDK\u540d\u79f0\u5217\u8868"), backup.get("\u6837\u672cSDK\u540d\u79f0\u5217\u8868")),
        "sdk_list": first_value(preferred.get("\u6837\u672cSDK\u540d\u79f0\u5217\u8868"), backup.get("\u6837\u672cSDK\u540d\u79f0\u5217\u8868")),
        "packer": "",
        "code_fuscator": "",
        "unshell_info": first_value(preferred.get("steady"), backup.get("steady")),
        "fake_app": normalize_bool(first_value(preferred.get("fakeApp"), backup.get("fakeApp"), preferred.get("\u662f\u5426\u4eff\u5192"), backup.get("\u662f\u5426\u4eff\u5192"))),
        "genuine": "true" if first_value(preferred.get("\u6b63\u7248\u5e94\u7528md5"), backup.get("\u6b63\u7248\u5e94\u7528md5")) else "",
        "official_app_name": first_value(preferred.get("\u6b63\u7248\u5e94\u7528\u540d\u79f0"), backup.get("\u6b63\u7248\u5e94\u7528\u540d\u79f0")),
        "rebuild_type": first_value(preferred.get("APP\u7248\u672c\u72b6\u6001\uff1a1\u3001\u6b63\u7248\uff0c2\u3001\u76d7\u7248\uff0c3\u3001\u53cd\u8bc8\u6807\u7b7e\u65e0\u503c"), backup.get("APP\u7248\u672c\u72b6\u6001\uff1a1\u3001\u6b63\u7248\uff0c2\u3001\u76d7\u7248\uff0c3\u3001\u53cd\u8bc8\u6807\u7b7e\u65e0\u503c")),
        "icon_md5": "",
        "virus_name": first_value(manual.get("engine_cm_virus_name"), manual.get("engine_360_virus_name"), preferred.get("\u75c5\u6bd2\u540d\u79f0"), backup.get("\u75c5\u6bd2\u540d\u79f0")),
        "risk_type": first_value(manual.get("engine_cm_type"), manual.get("engine_360_type"), preferred.get("\u7c7b\u578b"), backup.get("\u7c7b\u578b")),
        "virus_description": clean_text(json.dumps(engine_summary, ensure_ascii=False)),
        "fraud_flag": normalize_bool(first_value(preferred.get("\u662f\u5426\u6d89\u8bc8\u5e94\u7528\uff1a1\u3001\u662f\uff0c2\u3001\u5426"), backup.get("\u662f\u5426\u6d89\u8bc8\u5e94\u7528\uff1a1\u3001\u662f\uff0c2\u3001\u5426"))),
        "fraud_type_info": first_value(preferred.get("\u53cd\u8bc8\u6807\u7b7e"), backup.get("\u53cd\u8bc8\u6807\u7b7e")),
        "fraud_name": first_value(preferred.get("\u6d89\u8bc8\u5e94\u7528\u5bb6\u65cf"), backup.get("\u6d89\u8bc8\u5e94\u7528\u5bb6\u65cf")),
        "fraud_category": first_value(preferred.get("appType"), backup.get("appType")),
        "fraud_category_big": first_value(preferred.get("\u6d89\u8bc8\u5e94\u7528\u5927\u7c7b"), backup.get("\u6d89\u8bc8\u5e94\u7528\u5927\u7c7b")),
        "fraud_category_small": first_value(preferred.get("\u6d89\u8bc8\u5e94\u7528\u5c0f\u7c7b"), backup.get("\u6d89\u8bc8\u5e94\u7528\u5c0f\u7c7b")),
        "control_url": first_value(preferred.get("controlUrl"), backup.get("controlUrl")),
        "download_url": first_value(preferred.get("downloadUrl"), backup.get("downloadUrl")),
        "lt_urls": "",
        "dynamic_nets": "",
        "domains": "",
        "top_domains": "",
        "ips": "",
        "countries": first_value(preferred.get("\u6765\u6e90\u5730\u533a"), backup.get("\u6765\u6e90\u5730\u533a")),
        "operators": first_value(preferred.get("\u8fd0\u8425\u5546"), backup.get("\u8fd0\u8425\u5546")),
        "domain_count": "",
        "ip_count": "",
        "network_record_count": "",
        "source_risk_score": score_number(manual.get("engine_360_score"), manual.get("engine_cm_score"), manual.get("manual_review_result")),
        "source_malicious": source_malicious,
        "source_violation": "manual_360_cm_conflict",
        "query_time": first_value(preferred.get("findtime"), backup.get("findtime"), preferred.get("\u66f4\u65b0\u65f6\u95f4"), backup.get("\u66f4\u65b0\u65f6\u95f4")),
        "gold_label": gold_label,
        "label_source": manual.get("label_source"),
        "manual_review_result": manual.get("manual_review_result"),
        "conflict_type": manual.get("conflict_type"),
        "engine_360_type": manual.get("engine_360_type"),
        "engine_360_score": manual.get("engine_360_score"),
        "engine_360_virus_name": manual.get("engine_360_virus_name"),
        "engine_cm_type": manual.get("engine_cm_type"),
        "engine_cm_score": manual.get("engine_cm_score"),
        "engine_cm_virus_name": manual.get("engine_cm_virus_name"),
    }


def write_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(OUTPUT_SHEET)
    headers = APP_HEADERS + EXTRA_HEADERS
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--engine-360", required=True)
    parser.add_argument("--engine-cm", required=True)
    parser.add_argument("--manual", nargs="+", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    manual = load_manual_labels([Path(item) for item in args.manual])
    wanted = set(manual)
    rows360 = load_engine_rows(Path(args.engine_360), wanted)
    rowscm = load_engine_rows(Path(args.engine_cm), wanted)
    rows = [convert_record(manual[md5], rows360.get(md5, {}), rowscm.get(md5, {})) for md5 in sorted(wanted)]
    output = Path(args.out)
    output.parent.mkdir(parents=True, exist_ok=True)
    write_workbook(output, rows)
    summary = {
        "manual_labeled_rows": len(manual),
        "joined_360": len(rows360),
        "joined_cm": len(rowscm),
        "output_rows": len(rows),
        "output": str(output),
    }
    (output.parent / "\u4eba\u5de5\u6807\u6ce8\u51b2\u7a81\u6837\u672c_\u8f6c\u6362\u6458\u8981.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
