from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import sqlite3
import statistics
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parent
DATA_ROOT = WORKSPACE / "数据"
MALICIOUS_DIR = ROOT / "output_app_judgment"
WHITE_DIR = ROOT / "output_app_white_lite"
OUTPUT_DIR = ROOT / "training_artifacts"
DB_PATH = OUTPUT_DIR / "training_dataset.db"
SCHEMA_PATH = Path(__file__).with_name("evidence_schema.json")

ENGINE_FILES = {"engine_a": DATA_ROOT / "360.xlsx", "engine_b": DATA_ROOT / "cm.xlsx"}
CONFLICT_FILE = DATA_ROOT / "冲突样本分析_人工标注_分身规则更新.xlsx"
INVISIBLE_RE = re.compile(r"[\ufeff\u200b\u200c\u200d\u2060\u200a]")
BATCH_RE = re.compile(r"第(\d+)批_共(\d+)批_(\d{8}_\d{6})")

FIELD_ALIASES = {
    "MD5": "md5",
    "md5": "md5",
    "应用名称": "app_name",
    "input_appName": "app_name",
    "病毒名称": "virus_name",
    "病毒描述": "virus_description",
    "类型": "engine_label_text",
    "score": "engine_score",
    "危害类型": "harm_type",
    "fakeApp": "fake_app",
    "steady": "packer",
    "platform": "platform",
    "size": "file_size",
    "controlUrl": "control_url",
    "controlUrlMd5": "control_url_md5",
    "downloadUrl": "download_url",
    "downloadUrlMd5": "download_url_md5",
    "controlMailbox": "control_mailbox",
    "controlPhone": "control_phone",
    "反诈标签": "anti_fraud_tag",
    "样本sha1": "sha1",
    "样本sha256": "sha256",
    "APP版本状态：1、正版，2、盗版，3、反诈标签无值": "app_version_status",
    "是否涉诈应用：1、是，2、否": "fraud_flag",
    "涉诈应用大类": "fraud_category_big",
    "涉诈应用小类": "fraud_category_small",
    "涉诈应用家族": "fraud_family",
    "是否仿冒": "impersonation_flag",
    "正版应用icon": "official_icon",
    "正版应用的pkg": "official_pkg",
    "正版应用名称": "official_app_name",
    "正版应用版本号": "official_version",
    "正版应用md5": "official_md5",
    "仿冒一级分类": "impersonation_level1",
    "仿冒二级分类": "impersonation_level2",
    "仿冒三级分类": "impersonation_level3",
    "样本SDK名称列表": "sdk_list",
    "安装包文件类型": "file_type",
    "证书指纹MD5": "certificate_fingerprint",
    "应用签名证书sha1": "cert_sha1",
    "应用签名证书sha256": "cert_sha256",
    "证书拥有者": "certificate_owner",
    "证书开发者": "certificate_developer",
    "开发者名称": "developer_name",
    "样本appId": "sample_app_id",
    "黑客组织名称": "hacker_group",
    "样本文件是否已删除": "file_deleted",
    "input_fraudGaType": "fraud_category_big",
    "input_fraudGaSubType": "fraud_category_small",
    "人工审核结果": "human_label",
    "冲突类型": "conflict_type",
    "应用名称_360": "engine_a_app_name",
    "类型_360": "engine_a_label_text",
    "病毒名称_360": "engine_a_virus_name",
    "score_360": "engine_a_score",
    "应用名称_cm": "engine_b_app_name",
    "类型_cm": "engine_b_label_text",
    "病毒名称_cm": "engine_b_virus_name",
    "score_cm": "engine_b_score",
}

AGENTS = ("static_analysis", "threat_intel", "impersonation", "business_label")


def clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        value = INVISIBLE_RE.sub("", value).strip()
        return "" if value.lower() in {"nan", "none", "null"} else value
    return value


def md5_value(value: Any) -> str:
    text = str(clean(value)).upper()
    return text if re.fullmatch(r"[A-F0-9]{32}", text) else ""


def iter_sheet(path: Path, sheet_name: str | None = None) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(clean(item) or f"column_{index + 1}") for index, item in enumerate(next(rows, ()))]
        for values in rows:
            row = {
                headers[index]: clean(value)
                for index, value in enumerate(values[: len(headers)])
                if clean(value) not in ("", None)
            }
            if row:
                yield row
    finally:
        workbook.close()


def normalized_row(raw: dict[str, Any]) -> dict[str, Any]:
    result = {}
    for key, value in raw.items():
        result[FIELD_ALIASES.get(key, key)] = clean(value)
    if result.get("md5"):
        result["md5"] = md5_value(result["md5"])
    return result


def init_db() -> sqlite3.Connection:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS engine_features (
            md5 TEXT NOT NULL,
            engine TEXT NOT NULL,
            payload_json TEXT NOT NULL,
            PRIMARY KEY (md5, engine)
        );
        CREATE TABLE IF NOT EXISTS labels (
            md5 TEXT NOT NULL,
            label INTEGER,
            label_source TEXT NOT NULL,
            label_priority INTEGER NOT NULL,
            app_name TEXT,
            fraud_category_big TEXT,
            fraud_category_small TEXT,
            conflict_type TEXT,
            raw_json TEXT NOT NULL,
            PRIMARY KEY (md5, label_source)
        );
        CREATE TABLE IF NOT EXISTS unified_samples (
            md5 TEXT PRIMARY KEY,
            label INTEGER,
            label_source TEXT,
            split TEXT NOT NULL,
            group_id TEXT NOT NULL,
            conflict_type TEXT,
            payload_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS evidence_blocks (
            md5 TEXT NOT NULL,
            agent TEXT NOT NULL,
            split TEXT NOT NULL,
            label INTEGER,
            block_json TEXT NOT NULL,
            feature_json TEXT NOT NULL,
            PRIMARY KEY (md5, agent)
        );
        """
    )
    return conn


def audit() -> dict[str, Any]:
    report: dict[str, Any] = {"files": [], "sample_directories": {}, "warnings": []}
    for path in [*ENGINE_FILES.values(), CONFLICT_FILE]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            report["files"].append(
                {
                    "path": str(path),
                    "size_bytes": path.stat().st_size,
                    "sheets": [
                        {"name": sheet.title, "rows": sheet.max_row, "columns": sheet.max_column}
                        for sheet in workbook.worksheets
                    ],
                }
            )
        finally:
            workbook.close()
    for name, directory in {"malicious": MALICIOUS_DIR, "white": WHITE_DIR}.items():
        all_files = sorted(directory.glob("*.xlsx"))
        files = selected_export_files(directory)
        report["sample_directories"][name] = {
            "path": str(directory),
            "all_files": [path.name for path in all_files],
            "selected_complete_export": [path.name for path in files],
            "selected_file_count": len(files),
            "selected_rows": sum(workbook_rows(path) for path in files),
            "excluded_snapshot_files": [
                path.name for path in all_files if path not in files
            ],
        }
    report["warnings"] = [
        "恶意/白样本目录只有MD5、应用名和业务标签；静态、情报、仿冒证据必须通过MD5关联360/cm特征。",
        "人工冲突表中部分人工审核结果为空，只能进入待标注集，不能用于监督训练。",
        "当前没有APK/DEX/SO和图标原文件，相关证据只能标记缺失。",
    ]
    write_json(OUTPUT_DIR / "audit_report.json", report)
    return report


def workbook_rows(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return max(0, workbook[workbook.sheetnames[0]].max_row - 1)
    finally:
        workbook.close()


def prepare() -> dict[str, Any]:
    conn = init_db()
    stats = Counter()
    try:
        conn.execute("DELETE FROM engine_features")
        conn.execute("DELETE FROM labels")
        conn.execute("DELETE FROM unified_samples")
        import_directory_labels(conn, MALICIOUS_DIR, 1, "known_malicious", 20, stats)
        import_directory_labels(conn, WHITE_DIR, 0, "known_white", 20, stats)
        import_manual_conflicts(conn, stats)
        target_md5s = {
            row[0]
            for row in conn.execute("SELECT DISTINCT md5 FROM labels")
        }
        stats["target_labeled_md5"] = len(target_md5s)
        progress(f"待关联的有标签样本：{len(target_md5s):,} 个 MD5")
        for engine, path in ENGINE_FILES.items():
            batch = []
            scanned = 0
            matched = 0
            progress(f"开始扫描 {engine}: {path.name}")
            for raw in iter_sheet(path):
                scanned += 1
                row = normalized_row(raw)
                md5 = md5_value(row.get("md5"))
                if not md5:
                    stats[f"{engine}_invalid_md5"] += 1
                    continue
                if md5 not in target_md5s:
                    if scanned % 20000 == 0:
                        progress(
                            f"{engine} 已扫描 {scanned:,} 行，命中 {matched:,} 个目标样本"
                        )
                    continue
                matched += 1
                row["engine"] = engine
                row[f"{engine}_score"] = numeric(row.pop("engine_score", ""))
                row[f"{engine}_label_text"] = row.pop("engine_label_text", "")
                row[f"{engine}_label"] = label_from_engine(
                    row.get(f"{engine}_label_text"), row.get(f"{engine}_score")
                )
                batch.append((md5, engine, json.dumps(row, ensure_ascii=False)))
                if len(batch) >= 1000:
                    conn.executemany(
                        "INSERT OR REPLACE INTO engine_features VALUES (?, ?, ?)", batch
                    )
                    conn.commit()
                    stats[f"{engine}_rows"] += len(batch)
                    batch.clear()
                if scanned % 20000 == 0:
                    progress(
                        f"{engine} 已扫描 {scanned:,} 行，命中 {matched:,} 个目标样本"
                    )
            if batch:
                conn.executemany("INSERT OR REPLACE INTO engine_features VALUES (?, ?, ?)", batch)
                conn.commit()
                stats[f"{engine}_rows"] += len(batch)
            stats[f"{engine}_scanned_rows"] = scanned
            progress(f"{engine} 扫描完成：{scanned:,} 行，命中 {matched:,} 个目标样本")
        build_unified_samples(conn, stats)
    finally:
        conn.close()
    result = dict(stats)
    write_json(OUTPUT_DIR / "prepare_report.json", result)
    return result


def import_directory_labels(
    conn: sqlite3.Connection,
    directory: Path,
    label: int,
    source: str,
    priority: int,
    stats: Counter,
) -> None:
    seen = set()
    selected_files = selected_export_files(directory)
    stats[f"{source}_selected_files"] = len(selected_files)
    progress(
        f"{source} 选择完整导出组："
        + "、".join(path.name for path in selected_files)
    )
    for path in selected_files:
        file_rows = 0
        for raw in iter_sheet(path):
            file_rows += 1
            row = normalized_row(raw)
            md5 = md5_value(row.get("md5"))
            if not md5 or md5 in seen:
                continue
            seen.add(md5)
            conn.execute(
                """
                INSERT OR REPLACE INTO labels
                (md5,label,label_source,label_priority,app_name,fraud_category_big,
                 fraud_category_small,conflict_type,raw_json)
                VALUES (?,?,?,?,?,?,?,?,?)
                """,
                (
                    md5,
                    label,
                    source,
                    priority,
                    row.get("app_name", ""),
                    row.get("fraud_category_big", ""),
                    row.get("fraud_category_small", ""),
                    "",
                    json.dumps(row, ensure_ascii=False),
                ),
            )
        stats[f"{source}_rows"] += file_rows
        progress(f"{source} 已读取 {path.name}: {file_rows:,} 行")
    conn.commit()
    stats[f"{source}_unique"] = len(seen)


def import_manual_conflicts(conn: sqlite3.Connection, stats: Counter) -> None:
    workbook = load_workbook(CONFLICT_FILE, read_only=True, data_only=True)
    sheet_names = [name for name in workbook.sheetnames if "冲突" in name and "汇总" not in name]
    workbook.close()
    seen = set()
    for sheet in sheet_names:
        sheet_rows = 0
        for raw in iter_sheet(CONFLICT_FILE, sheet):
            sheet_rows += 1
            row = normalized_row(raw)
            md5 = md5_value(row.get("md5"))
            if not md5 or md5 in seen:
                continue
            seen.add(md5)
            label = manual_label(row.get("human_label"))
            conn.execute(
                """
                INSERT OR REPLACE INTO labels
                (md5,label,label_source,label_priority,app_name,fraud_category_big,
                 fraud_category_small,conflict_type,raw_json)
                VALUES (?,?,?,?,?,?,?,?,?)
                """,
                (
                    md5,
                    label,
                    "manual_conflict",
                    100,
                    first(row, "engine_b_app_name", "engine_a_app_name"),
                    "",
                    "",
                    row.get("conflict_type", sheet),
                    json.dumps(row, ensure_ascii=False),
                ),
            )
            stats["manual_conflict_labeled" if label is not None else "manual_conflict_unlabeled"] += 1
        progress(f"人工冲突表 {sheet}: 读取 {sheet_rows:,} 行")
    conn.commit()


def build_unified_samples(conn: sqlite3.Connection, stats: Counter) -> None:
    md5s = [
        row[0]
        for row in conn.execute(
            "SELECT DISTINCT md5 FROM labels"
        )
    ]
    for md5 in md5s:
        payload: dict[str, Any] = {"md5": md5}
        for row in conn.execute(
            "SELECT engine,payload_json FROM engine_features WHERE md5=?", (md5,)
        ):
            engine_payload = json.loads(row[1])
            payload.update(
                {
                    key: value
                    for key, value in engine_payload.items()
                    if value not in ("", None, [], {})
                }
            )
        labels = conn.execute(
            """
            SELECT label,label_source,label_priority,app_name,fraud_category_big,
                   fraud_category_small,conflict_type,raw_json
            FROM labels WHERE md5=? ORDER BY label_priority DESC
            """,
            (md5,),
        ).fetchall()
        manual = next(
            (row for row in labels if row[1] == "manual_conflict" and row[0] is not None),
            None,
        )
        known_values = {
            row[0]
            for row in labels
            if row[1] in {"known_malicious", "known_white"} and row[0] is not None
        }
        if manual:
            chosen = manual
        elif len(known_values) > 1:
            chosen = None
            label, source, conflict = None, "known_label_conflict", "恶意/白样本标签冲突"
            stats["excluded_known_label_conflict"] += 1
        else:
            chosen = next(
                (row for row in labels if row[0] is not None),
                labels[0] if labels else None,
            )
        if chosen:
            label, source, _, app_name, big, small, conflict, raw_json = chosen
            payload.update(
                {
                    key: value
                    for key, value in json.loads(raw_json).items()
                    if value not in ("", None, [], {})
                }
            )
            payload["app_name"] = app_name or payload.get("app_name", "")
            payload["fraud_category_big"] = big or payload.get("fraud_category_big", "")
            payload["fraud_category_small"] = small or payload.get("fraud_category_small", "")
        elif not labels:
            label, source, conflict = None, "unlabeled_engine", ""
        derive_fields(payload)
        group_id = sample_group(payload)
        split = stable_split(group_id)
        conn.execute(
            "INSERT OR REPLACE INTO unified_samples VALUES (?,?,?,?,?,?,?)",
            (
                md5,
                label,
                source,
                split,
                group_id,
                conflict or payload.get("conflict_type", ""),
                json.dumps(payload, ensure_ascii=False),
            ),
        )
        stats[f"unified_{split}"] += 1
        stats["unified_labeled" if label is not None else "unified_unlabeled"] += 1
    conn.commit()


def derive_fields(payload: dict[str, Any]) -> None:
    for engine in ("engine_a", "engine_b"):
        score = numeric(payload.get(f"{engine}_score"))
        if score is not None:
            payload[f"{engine}_score"] = score
        payload.setdefault(
            f"{engine}_label",
            label_from_engine(payload.get(f"{engine}_label_text"), score),
        )
    payload["app_name"] = first(
        payload,
        "app_name",
        "engine_b_app_name",
        "engine_a_app_name",
    )
    payload["virus_name"] = first(
        payload,
        "virus_name",
        "engine_b_virus_name",
        "engine_a_virus_name",
    )
    a, b = numeric(payload.get("engine_a_score")), numeric(payload.get("engine_b_score"))
    if a is not None and b is not None:
        payload["engine_score_gap"] = abs(a - b)
        payload["engine_disagreement"] = float(
            payload.get("engine_a_label") != payload.get("engine_b_label")
        )


def generate_sft_and_freeze() -> dict[str, Any]:
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    conn = init_db()
    conn.execute("DELETE FROM evidence_blocks")
    paths = {
        (agent, split): (OUTPUT_DIR / "sft" / agent / f"{split}.jsonl")
        for agent in AGENTS
        for split in ("train", "val", "test")
    }
    for path in paths.values():
        path.parent.mkdir(parents=True, exist_ok=True)
    frozen = OUTPUT_DIR / "frozen_evidence_blocks.jsonl"
    counts = Counter()
    streams = {
        key: path.open("w", encoding="utf-8", newline="\n")
        for key, path in paths.items()
    }
    frozen_stream = frozen.open("w", encoding="utf-8", newline="\n")
    try:
        rows = conn.execute(
            "SELECT md5,label,split,payload_json FROM unified_samples WHERE label IS NOT NULL"
        )
        total = 0
        for md5, label, split, payload_json in rows:
            total += 1
            sample = json.loads(payload_json)
            blocks = []
            for agent in AGENTS:
                block, features = evidence_block(agent, sample, schema)
                blocks.append(block)
                conn.execute(
                    "INSERT OR REPLACE INTO evidence_blocks VALUES (?,?,?,?,?,?)",
                    (
                        md5,
                        agent,
                        split,
                        label,
                        json.dumps(block, ensure_ascii=False),
                        json.dumps(features, ensure_ascii=False),
                    ),
                )
                record = {
                    "messages": [
                        {
                            "role": "system",
                            "content": (
                                f"你是{agent}领域智能体。只能引用输入字段，"
                                "严格输出证据Schema JSON，不得编造缺失证据。"
                            ),
                        },
                        {
                            "role": "user",
                            "content": json.dumps(agent_input(agent, sample, schema), ensure_ascii=False),
                        },
                        {
                            "role": "assistant",
                            "content": json.dumps(block, ensure_ascii=False),
                        },
                    ],
                    "metadata": {
                        "md5": md5,
                        "label": label,
                        "split": split,
                        "evidence_label_type": "silver_rule_generated",
                    },
                }
                streams[(agent, split)].write(
                    json.dumps(record, ensure_ascii=False) + "\n"
                )
                counts[f"{agent}_{split}"] += 1
            frozen_stream.write(
                json.dumps(
                    {
                        "md5": md5,
                        "label": label,
                        "split": split,
                        "evidence_blocks": blocks,
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )
            if total % 2000 == 0:
                conn.commit()
                progress(f"已生成并冻结 {total:,} 个样本的四智能体证据")
        conn.commit()
    finally:
        for stream in streams.values():
            stream.close()
        frozen_stream.close()
        conn.close()
    result = dict(counts)
    result["warning"] = "当前证据为Schema规则生成的银标，SFT前建议抽样人工复核。"
    write_json(OUTPUT_DIR / "sft_freeze_report.json", result)
    return result


def evidence_block(
    agent: str, sample: dict[str, Any], schema: dict[str, Any]
) -> tuple[dict[str, Any], dict[str, float]]:
    spec = schema["agents"][agent]
    evidence = []
    features: dict[str, float] = {}
    for evidence_type, evidence_spec in spec["evidence_types"].items():
        item = evidence_item(agent, evidence_type, evidence_spec, sample)
        features[evidence_type] = item["strength"] if item else 0.0
        if item:
            evidence.append(item)
    missing = []
    present_groups = 0
    for fields in spec["required_field_groups"]:
        if any(present(sample.get(field)) for field in fields):
            present_groups += 1
        else:
            missing.append("/".join(fields))
    evidence_strength = min(1.0, sum(item["strength"] for item in evidence) / max(1, len(evidence)))
    completeness = present_groups / max(1, len(spec["required_field_groups"]))
    risk_score = min(1.0, sum(features.values()) / max(1.0, len(features) * 0.55))
    confidence = round(0.25 + 0.45 * completeness + 0.30 * evidence_strength, 4)
    features["evidence_count"] = float(len(evidence))
    features["completeness"] = completeness
    return (
        {
            "agent": agent,
            "claim": risk_claim(agent, risk_score),
            "risk_score": round(risk_score, 4),
            "confidence": min(0.98, confidence),
            "evidence_strength": round(evidence_strength, 4),
            "evidence": evidence,
            "missing_fields": missing,
        },
        features,
    )


def evidence_item(
    agent: str,
    evidence_type: str,
    spec: dict[str, Any],
    sample: dict[str, Any],
) -> dict[str, Any] | None:
    values = [(field, sample.get(field)) for field in spec["fields"] if present(sample.get(field))]
    if not values:
        return None
    strength, direction = evidence_strength(agent, evidence_type, values, sample)
    if strength <= 0:
        return None
    return {
        "evidence_type": evidence_type,
        "source_fields": [field for field, _ in values],
        "source_values": [safe_value(value) for _, value in values],
        "direction": direction,
        "strength": round(strength, 4),
        "description": spec["description"],
    }


def evidence_strength(
    agent: str,
    evidence_type: str,
    values: list[tuple[str, Any]],
    sample: dict[str, Any],
) -> tuple[float, str]:
    text = " ".join(str(value).lower() for _, value in values)
    if evidence_type == "engine_static_conflict":
        gap = float(sample.get("engine_score_gap") or 0)
        return min(1.0, 0.25 + gap / 140), "context"
    if evidence_type == "signature_anomaly":
        suspicious = any(term in text for term in ("invalid", "tamper", "mismatch", "missing", "异常"))
        return (0.85, "supports_malicious") if suspicious else (0.25, "supports_benign")
    if evidence_type == "packer_or_obfuscation":
        return (0.65, "supports_malicious") if truthy(text) else (0.15, "context")
    if evidence_type in {"control_infrastructure", "download_infrastructure"}:
        return 0.72, "supports_malicious"
    if evidence_type == "malware_family":
        malicious = any(term in text for term in ("fraud", "trojan", "rogue", "spy", "phish", "木马", "病毒"))
        return (0.82 if malicious else 0.35), ("supports_malicious" if malicious else "context")
    if evidence_type in {"fraud_intelligence", "fraud_category", "anti_fraud_label"}:
        benign = any(term in text for term in ("否", "安全应用", "白样本"))
        return (0.22, "supports_benign") if benign else (0.8, "supports_malicious")
    if evidence_type == "declared_impersonation":
        positive = any(term in text for term in ("是", "盗版", "仿冒", "true", "1"))
        return (0.88 if positive else 0.2), ("supports_malicious" if positive else "supports_benign")
    if evidence_type == "official_asset_reference":
        return 0.55, "context"
    if evidence_type == "impersonation_category":
        return 0.78, "supports_malicious"
    if evidence_type == "name_obfuscation":
        name = str(sample.get("app_name") or "")
        obfuscated = bool(re.search(r"\d{2,}$", name)) or bool(INVISIBLE_RE.search(name))
        return (0.62, "supports_malicious") if obfuscated else (0.15, "context")
    if evidence_type == "app_name_semantics":
        risky = any(
            term in text
            for term in ("贷款", "博彩", "赌博", "色情", "刷单", "投资", "返利", "约炮", "棋牌")
        )
        return (0.65 if risky else 0.18), ("supports_malicious" if risky else "context")
    if evidence_type == "harm_category":
        return 0.7, "supports_malicious"
    return min(0.65, 0.2 + 0.12 * len(values)), "context"


def train_models() -> dict[str, Any]:
    conn = init_db()
    model_dir = OUTPUT_DIR / "models"
    model_dir.mkdir(parents=True, exist_ok=True)
    agent_models = {}
    report: dict[str, Any] = {"agents": {}}
    try:
        for agent in AGENTS:
            rows = load_agent_rows(conn, agent)
            model, metrics = fit_binary_model(rows, feature_names_for_rows(rows))
            agent_models[agent] = model
            write_json(model_dir / f"{agent}.json", model)
            report["agents"][agent] = metrics

        stacked = build_stacked_rows(conn, agent_models)
        fusion_features = [f"{agent}_prob" for agent in AGENTS] + [
            f"{agent}_confidence" for agent in AGENTS
        ]
        fusion, fusion_metrics = fit_binary_model(stacked, fusion_features)
        write_json(model_dir / "fusion.json", fusion)
        report["fusion"] = fusion_metrics

        wec_rows = build_wec_rows(stacked, fusion)
        wec_features = [
            "engine_a_prob",
            "engine_b_prob",
            "engine_c_prob",
            "engine_score_gap",
            "engine_disagreement",
        ]
        wec, wec_metrics = fit_binary_model(wec_rows, wec_features)
        write_json(model_dir / "wec.json", wec)
        report["wec"] = wec_metrics

        arbiter_rows = [row for row in wec_rows if row.get("label_source") == "manual_conflict"]
        if class_count(arbiter_rows) >= 2 and len(arbiter_rows) >= 20:
            arbiter, arbiter_metrics = fit_binary_model(arbiter_rows, wec_features)
        else:
            arbiter, arbiter_metrics = wec, {
                "warning": "人工冲突标注不足或只有单一类别，终审器暂时复用WEC模型。"
            }
        write_json(model_dir / "arbiter.json", arbiter)
        report["arbiter"] = arbiter_metrics
    finally:
        conn.close()
    write_json(OUTPUT_DIR / "training_report.json", report)
    return report


def calibrate() -> dict[str, Any]:
    conn = init_db()
    model_dir = OUTPUT_DIR / "models"
    report = {}
    try:
        for name in [*AGENTS, "fusion", "wec", "arbiter"]:
            path = model_dir / f"{name}.json"
            calibration_path = model_dir / f"{name}_calibration.json"
            if not path.exists():
                continue
            model = json.loads(path.read_text(encoding="utf-8"))
            rows = calibration_rows(conn, name, model, model_dir)
            if not rows:
                calibration_path.unlink(missing_ok=True)
                report[name] = {"warning": "没有验证集数据"}
                continue
            probabilities = [row["probability"] for row in rows]
            labels = [row["label"] for row in rows]
            if len(set(labels)) < 2:
                calibration_path.unlink(missing_ok=True)
                report[name] = {
                    "warning": "验证集只有单一类别，拒绝生成概率校准器。",
                    "validation_count": len(labels),
                    "classes": sorted(set(labels)),
                }
                continue
            temperature = fit_temperature(probabilities, labels)
            isotonic = fit_isotonic(probabilities, labels)
            raw_brier = brier(probabilities, labels)
            temp_probs = [apply_temperature(p, temperature) for p in probabilities]
            iso_probs = [apply_isotonic(p, isotonic) for p in probabilities]
            unique_probability_count = len(
                {round(float(value), 12) for value in probabilities}
            )
            candidates = {
                "raw": raw_brier,
                "temperature": brier(temp_probs, labels),
            }
            if unique_probability_count >= 10:
                candidates["isotonic"] = brier(iso_probs, labels)
            selected = min(candidates, key=candidates.get)
            calibration = {
                "selected": selected,
                "validation_count": len(labels),
                "unique_probability_count": unique_probability_count,
                "temperature": temperature,
                "isotonic": isotonic,
                "validation_brier": candidates,
                "validation_ece": {
                    "raw": ece(probabilities, labels),
                    "temperature": ece(temp_probs, labels),
                    "isotonic": ece(iso_probs, labels),
                },
            }
            write_json(calibration_path, calibration)
            report[name] = calibration
    finally:
        conn.close()
    write_json(OUTPUT_DIR / "calibration_report.json", report)
    return report


def load_agent_rows(conn: sqlite3.Connection, agent: str) -> list[dict[str, Any]]:
    result = []
    for md5, label, split, feature_json in conn.execute(
        "SELECT md5,label,split,feature_json FROM evidence_blocks WHERE agent=? AND label IS NOT NULL",
        (agent,),
    ):
        features = json.loads(feature_json)
        result.append({"md5": md5, "label": int(label), "split": split, **features})
    return result


def feature_names_for_rows(rows: list[dict[str, Any]]) -> list[str]:
    ignored = {"md5", "label", "split", "label_source"}
    return sorted({key for row in rows for key in row if key not in ignored})


def fit_binary_model(
    rows: list[dict[str, Any]], feature_names: list[str]
) -> tuple[dict[str, Any], dict[str, Any]]:
    train = [row for row in rows if row["split"] == "train"]
    val = [row for row in rows if row["split"] == "val"]
    test = [row for row in rows if row["split"] == "test"]
    if not train or class_count(train) < 2:
        model = constant_model(rows, feature_names)
        return model, {"warning": "训练数据不足或只有单一类别，使用常量模型。"}
    x = np.array([[float(row.get(name, 0.0)) for name in feature_names] for row in train])
    y = np.array([float(row["label"]) for row in train])
    mean = x.mean(axis=0)
    scale = x.std(axis=0)
    scale[scale < 1e-8] = 1.0
    z = (x - mean) / scale
    weights = np.zeros(z.shape[1])
    bias = logit(float(np.clip(y.mean(), 1e-4, 1 - 1e-4)))
    positive_count = max(1.0, float(y.sum()))
    negative_count = max(1.0, float(len(y) - y.sum()))
    sample_weights = np.where(
        y == 1,
        len(y) / (2.0 * positive_count),
        len(y) / (2.0 * negative_count),
    )
    lr, l2 = 0.08, 0.01
    for _ in range(700):
        pred = sigmoid_array(z @ weights + bias)
        error = (pred - y) * sample_weights
        weights -= lr * ((z.T @ error) / len(y) + l2 * weights)
        bias -= lr * float(error.mean())
    model = {
        "type": "standardized_logistic_regression",
        "feature_names": feature_names,
        "mean": mean.tolist(),
        "scale": scale.tolist(),
        "weights": weights.tolist(),
        "bias": bias,
        "training": {
            "loss": "class_balanced_binary_cross_entropy",
            "l2": l2,
            "iterations": 700,
            "learning_rate": lr,
        },
    }
    return model, {
        "train": evaluate_model(model, train),
        "validation": evaluate_model(model, val),
        "test": evaluate_model(model, test),
        "top_weights": sorted(
            zip(feature_names, weights.tolist()), key=lambda item: abs(item[1]), reverse=True
        )[:20],
    }


def build_stacked_rows(
    conn: sqlite3.Connection, agent_models: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    by_md5: dict[str, dict[str, Any]] = {}
    query = """
        SELECT e.md5,e.agent,e.split,e.label,e.block_json,e.feature_json,
               u.label_source,u.payload_json
        FROM evidence_blocks e JOIN unified_samples u ON u.md5=e.md5
        WHERE e.label IS NOT NULL
    """
    for md5, agent, split, label, block_json, feature_json, source, payload_json in conn.execute(query):
        row = by_md5.setdefault(
            md5,
            {
                "md5": md5,
                "split": split,
                "label": int(label),
                "label_source": source,
                "sample": json.loads(payload_json),
            },
        )
        features = json.loads(feature_json)
        block = json.loads(block_json)
        row[f"{agent}_prob"] = predict(agent_models[agent], features)
        row[f"{agent}_confidence"] = float(block["confidence"])
    return list(by_md5.values())


def build_wec_rows(
    stacked: list[dict[str, Any]], fusion: dict[str, Any]
) -> list[dict[str, Any]]:
    result = []
    for row in stacked:
        sample = row.pop("sample")
        a = score_probability(sample.get("engine_a_score"), sample.get("engine_a_label"))
        b = score_probability(sample.get("engine_b_score"), sample.get("engine_b_label"))
        result.append(
            {
                **row,
                "engine_a_prob": a,
                "engine_b_prob": b,
                "engine_c_prob": predict(fusion, row),
                "engine_score_gap": abs(a - b),
                "engine_disagreement": float(
                    sample.get("engine_a_label") != sample.get("engine_b_label")
                ),
            }
        )
    return result


def calibration_rows(
    conn: sqlite3.Connection,
    name: str,
    model: dict[str, Any],
    model_dir: Path,
) -> list[dict[str, Any]]:
    if name in AGENTS:
        return [
            {**row, "probability": predict(model, row)}
            for row in load_agent_rows(conn, name)
            if row["split"] == "val"
        ]
    agent_models = {
        agent: json.loads((model_dir / f"{agent}.json").read_text(encoding="utf-8"))
        for agent in AGENTS
    }
    stacked = build_stacked_rows(conn, agent_models)
    if name == "fusion":
        rows = stacked
    else:
        fusion = json.loads((model_dir / "fusion.json").read_text(encoding="utf-8"))
        rows = build_wec_rows(stacked, fusion)
        if name == "arbiter":
            rows = [row for row in rows if row.get("label_source") == "manual_conflict"]
    return [
        {**row, "probability": predict(model, row)}
        for row in rows
        if row["split"] == "val"
    ]


def predict(model: dict[str, Any], row: dict[str, Any]) -> float:
    if model["type"] == "constant":
        return float(model["probability"])
    values = np.array([float(row.get(name, 0.0)) for name in model["feature_names"]])
    z = (values - np.array(model["mean"])) / np.array(model["scale"])
    return sigmoid(float(z @ np.array(model["weights"]) + model["bias"]))


def evaluate_model(model: dict[str, Any], rows: list[dict[str, Any]]) -> dict[str, Any]:
    if not rows:
        return {"count": 0}
    labels = [int(row["label"]) for row in rows]
    probs = [predict(model, row) for row in rows]
    preds = [int(value >= 0.5) for value in probs]
    tp = sum(p == y == 1 for p, y in zip(preds, labels))
    tn = sum(p == y == 0 for p, y in zip(preds, labels))
    fp = sum(p == 1 and y == 0 for p, y in zip(preds, labels))
    fn = sum(p == 0 and y == 1 for p, y in zip(preds, labels))
    precision = tp / max(1, tp + fp)
    recall = tp / max(1, tp + fn)
    specificity = tn / max(1, tn + fp)
    return {
        "count": len(rows),
        "positive": sum(labels),
        "accuracy": round((tp + tn) / len(rows), 4),
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "specificity": round(specificity, 4),
        "balanced_accuracy": round((recall + specificity) / 2, 4),
        "f1": round(2 * precision * recall / max(1e-12, precision + recall), 4),
        "roc_auc": round(roc_auc(probs, labels), 4),
        "brier": round(brier(probs, labels), 6),
        "ece": round(ece(probs, labels), 6),
    }


def fit_temperature(probabilities: list[float], labels: list[int]) -> float:
    best = (float("inf"), 1.0)
    logits = [logit(float(np.clip(p, 1e-6, 1 - 1e-6))) for p in probabilities]
    for temperature in np.linspace(0.25, 5.0, 191):
        probs = [sigmoid(value / float(temperature)) for value in logits]
        loss = log_loss(probs, labels)
        if loss < best[0]:
            best = (loss, float(temperature))
    return round(best[1], 4)


def fit_isotonic(probabilities: list[float], labels: list[int]) -> list[dict[str, float]]:
    grouped: dict[float, list[float]] = defaultdict(lambda: [0.0, 0.0])
    for probability, label in zip(probabilities, labels):
        key = round(float(probability), 12)
        grouped[key][0] += float(label)
        grouped[key][1] += 1.0
    blocks = [
        {"lower": p, "upper": p, "sum": values[0], "count": values[1]}
        for p, values in sorted(grouped.items())
    ]
    index = 0
    while index < len(blocks) - 1:
        left = blocks[index]["sum"] / blocks[index]["count"]
        right = blocks[index + 1]["sum"] / blocks[index + 1]["count"]
        if left <= right:
            index += 1
            continue
        merged = {
            "lower": blocks[index]["lower"],
            "upper": blocks[index + 1]["upper"],
            "sum": blocks[index]["sum"] + blocks[index + 1]["sum"],
            "count": blocks[index]["count"] + blocks[index + 1]["count"],
        }
        blocks[index : index + 2] = [merged]
        index = max(0, index - 1)
    return [
        {
            "lower": block["lower"],
            "upper": block["upper"],
            "value": round(block["sum"] / block["count"], 6),
        }
        for block in blocks
    ]


def apply_temperature(probability: float, temperature: float) -> float:
    return sigmoid(logit(float(np.clip(probability, 1e-6, 1 - 1e-6))) / temperature)


def apply_isotonic(probability: float, blocks: list[dict[str, float]]) -> float:
    if not blocks:
        return probability
    for block in blocks:
        if probability <= block["upper"]:
            return block["value"]
    return blocks[-1]["value"]


def class_count(rows: list[dict[str, Any]]) -> int:
    return len({row["label"] for row in rows})


def constant_model(rows: list[dict[str, Any]], features: list[str]) -> dict[str, Any]:
    probability = statistics.fmean([row["label"] for row in rows]) if rows else 0.5
    return {"type": "constant", "feature_names": features, "probability": probability}


def label_from_engine(text: Any, score: Any) -> str:
    value = str(text or "").lower()
    if any(term in value for term in ("恶意", "病毒", "木马")):
        return "malicious"
    if "疑似" in value or "可疑" in value:
        return "suspicious"
    if "白" in value or "良性" in value:
        return "benign"
    number = numeric(score)
    if number is None:
        return ""
    return "malicious" if number >= 90 else "suspicious" if number >= 45 else "benign"


def manual_label(value: Any) -> int | None:
    if value in ("", None):
        return None
    text = str(value).strip().lower()
    if text in {"100", "1", "恶意", "malicious", "是"}:
        return 1
    if text in {"0", "良性", "白", "benign", "否"}:
        return 0
    try:
        return int(float(text) >= 50)
    except ValueError:
        return None


def sample_group(sample: dict[str, Any]) -> str:
    app = re.sub(r"\d+$", "", str(sample.get("app_name") or "").lower())
    cert = first(sample, "cert_sha256", "cert_sha1", "certificate_fingerprint")
    family = first(sample, "fraud_family", "virus_name")
    seed = "|".join(item for item in (cert, family, app) if item) or sample["md5"]
    return hashlib.sha1(seed.encode("utf-8")).hexdigest()[:20]


def stable_split(group_id: str) -> str:
    bucket = int(hashlib.sha1(group_id.encode()).hexdigest()[:8], 16) % 100
    return "train" if bucket < 70 else "val" if bucket < 85 else "test"


def selected_export_files(directory: Path) -> list[Path]:
    """Select one complete export generation and exclude older snapshots."""
    groups: dict[str, list[tuple[int, int, Path]]] = defaultdict(list)
    unmatched = []
    for path in sorted(directory.glob("*.xlsx")):
        match = BATCH_RE.search(path.stem)
        if not match:
            unmatched.append(path)
            continue
        groups[match.group(3)].append((int(match.group(1)), int(match.group(2)), path))
    complete = []
    for export_id, items in groups.items():
        expected = max(item[1] for item in items)
        observed = sorted(item[0] for item in items)
        if observed == list(range(1, expected + 1)):
            complete.append((export_id, expected, [item[2] for item in sorted(items)]))
    if complete:
        # Prefer the largest complete generation, then the latest timestamp.
        return max(complete, key=lambda item: (item[1], item[0]))[2]
    if groups:
        latest = max(groups)
        return [item[2] for item in sorted(groups[latest])]
    return unmatched


def progress(message: str) -> None:
    print(f"[pipeline] {message}", file=sys.stderr, flush=True)


def score_probability(score: Any, label: Any) -> float:
    number = numeric(score)
    if number is not None:
        return float(np.clip(number / 100.0, 0, 1))
    return {"malicious": 0.9, "suspicious": 0.65, "benign": 0.1}.get(str(label), 0.5)


def agent_input(agent: str, sample: dict[str, Any], schema: dict[str, Any]) -> dict[str, Any]:
    fields = sorted(
        {
            field
            for item in schema["agents"][agent]["evidence_types"].values()
            for field in item["fields"]
        }
    )
    return {field: sample.get(field) for field in fields if present(sample.get(field))}


def risk_claim(agent: str, score: float) -> str:
    names = {
        "static_analysis": "静态特征",
        "threat_intel": "情报关联",
        "impersonation": "仿冒",
        "business_label": "业务影响",
    }
    level = "较高" if score >= 0.7 else "中等" if score >= 0.45 else "较低"
    return f"{names[agent]}风险{level}。"


def present(value: Any) -> bool:
    return clean(value) not in ("", None, [], {})


def truthy(value: Any) -> bool:
    return str(value).lower() in {"1", "true", "yes", "是", "仿冒", "盗版", "已删除"}


def numeric(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def first(mapping: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = clean(mapping.get(key))
        if value not in ("", None):
            return str(value)
    return ""


def safe_value(value: Any) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)[:500]
    return str(value)[:500]


def sigmoid(value: float) -> float:
    if value >= 0:
        z = math.exp(-value)
        return 1 / (1 + z)
    z = math.exp(value)
    return z / (1 + z)


def sigmoid_array(values: np.ndarray) -> np.ndarray:
    return np.where(values >= 0, 1 / (1 + np.exp(-values)), np.exp(values) / (1 + np.exp(values)))


def logit(value: float) -> float:
    return math.log(value / (1 - value))


def brier(probabilities: list[float], labels: list[int]) -> float:
    return statistics.fmean((p - y) ** 2 for p, y in zip(probabilities, labels))


def log_loss(probabilities: list[float], labels: list[int]) -> float:
    return statistics.fmean(
        -(y * math.log(max(p, 1e-9)) + (1 - y) * math.log(max(1 - p, 1e-9)))
        for p, y in zip(probabilities, labels)
    )


def roc_auc(probabilities: list[float], labels: list[int]) -> float:
    positives = sum(labels)
    negatives = len(labels) - positives
    if positives == 0 or negatives == 0:
        return 0.5
    ordered = sorted(zip(probabilities, labels), key=lambda item: item[0])
    rank_sum = 0.0
    index = 0
    while index < len(ordered):
        end = index + 1
        while end < len(ordered) and ordered[end][0] == ordered[index][0]:
            end += 1
        average_rank = (index + 1 + end) / 2.0
        rank_sum += average_rank * sum(label for _, label in ordered[index:end])
        index = end
    return (rank_sum - positives * (positives + 1) / 2) / (positives * negatives)


def ece(probabilities: list[float], labels: list[int], bins: int = 10) -> float:
    total = len(labels)
    error = 0.0
    for index in range(bins):
        lower, upper = index / bins, (index + 1) / bins
        selected = [
            (p, y)
            for p, y in zip(probabilities, labels)
            if lower <= p < upper or (index == bins - 1 and p == 1)
        ]
        if selected:
            confidence = statistics.fmean(item[0] for item in selected)
            accuracy = statistics.fmean(item[1] for item in selected)
            error += len(selected) / total * abs(confidence - accuracy)
    return error


def append_jsonl(path: Path, payload: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def run_all() -> dict[str, Any]:
    return {
        "audit": audit(),
        "prepare": prepare(),
        "sft_and_freeze": generate_sft_and_freeze(),
        "training": train_models(),
        "calibration": calibrate(),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="MalApp six-stage training pipeline")
    parser.add_argument(
        "command",
        choices=("audit", "prepare", "sft", "freeze", "train", "calibrate", "all"),
    )
    args = parser.parse_args()
    action = {
        "audit": audit,
        "prepare": prepare,
        "sft": generate_sft_and_freeze,
        "freeze": generate_sft_and_freeze,
        "train": train_models,
        "calibrate": calibrate,
        "all": run_all,
    }[args.command]
    print(json.dumps(action(), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
