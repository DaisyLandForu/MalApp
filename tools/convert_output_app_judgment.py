from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


PROFILE_SHEET = "\u63a5\u53e31_APP\u4e3b\u753b\u50cf"
PROFILE_LITE_SHEET = "\u63a5\u53e31_APP\u4e3b\u753b\u50cf_\u7cbe\u7b80"
NETWORK_SUMMARY_SHEET = "MD5\u901a\u8054\u6c47\u603b"
NETWORK_SUMMARY_LITE_SHEET = "MD5\u901a\u8054\u6c47\u603b_\u7cbe\u7b80"
RAW_INPUT_SHEET = "\u539f\u59cb\u8f93\u5165"
OUTPUT_SHEET = "APP\u7814\u5224\u6570\u636e"
SUMMARY_NAME = "\u8f6c\u6362\u6458\u8981.json"
DEFAULT_COMBINED_NAME = "output_app_judgment_APP\u53ef\u5bfc\u5165_\u5408\u5e76\u7248.xlsx"
SINGLE_SUFFIX = "_APP\u53ef\u5bfc\u5165\u5355Sheet.xlsx"


APP_HEADERS = [
    "md5",
    "sample_id",
    "app_name",
    "package_name",
    "sha1",
    "sha256",
    "version_name",
    "version_code",
    "file_size",
    "file_type",
    "signature_status",
    "certificate_fingerprint",
    "cert_sha1",
    "cert_sha256",
    "certificate_owner",
    "certificate_developer",
    "permissions",
    "plugins",
    "sdk_list",
    "packer",
    "code_fuscator",
    "unshell_info",
    "fake_app",
    "genuine",
    "official_app_name",
    "rebuild_type",
    "icon_md5",
    "virus_name",
    "risk_type",
    "virus_description",
    "fraud_flag",
    "fraud_type_info",
    "fraud_name",
    "fraud_category",
    "fraud_category_big",
    "fraud_category_small",
    "control_url",
    "download_url",
    "lt_urls",
    "dynamic_nets",
    "domains",
    "top_domains",
    "ips",
    "countries",
    "operators",
    "domain_count",
    "ip_count",
    "network_record_count",
    "source_risk_score",
    "source_malicious",
    "source_violation",
    "query_time",
]


FEATURE_HEADERS = [
    "package_name",
    "sha1",
    "sha256",
    "version_name",
    "file_size",
    "signature_status",
    "certificate_fingerprint",
    "permissions",
    "plugins",
    "sdk_list",
    "packer",
    "code_fuscator",
    "unshell_info",
    "fake_app",
    "genuine",
    "official_app_name",
    "rebuild_type",
    "icon_md5",
    "virus_name",
    "risk_type",
    "virus_description",
    "fraud_type_info",
    "control_url",
    "dynamic_nets",
    "domains",
    "top_domains",
    "ips",
    "countries",
    "operators",
    "domain_count",
    "ip_count",
    "network_record_count",
]


def value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        item = row.get(key)
        if item not in (None, ""):
            return item
    return ""


def clean_md5(item: Any) -> str:
    text = str(item or "").strip().upper()
    return text if re.fullmatch(r"[A-F0-9]{32}", text) else text


def normalize_bool(item: Any) -> str:
    if item in (None, ""):
        return ""
    text = str(item).strip().lower()
    if text in {"1", "true", "yes", "y", "\u662f"}:
        return "true"
    if text in {"0", "false", "no", "n", "\u5426"}:
        return "false"
    return str(item)


def normalize_signature_status(row: dict[str, Any]) -> str:
    signature = value(row, "signature", "sign", "certificate")
    if signature in (None, ""):
        return ""
    text = str(signature).strip().lower()
    if text in {"0", "false", "none", "null", "\u65e0"}:
        return "missing"
    return "present"


def compact_list_text(item: Any, max_chars: int = 12000) -> str:
    if item in (None, ""):
        return ""
    if isinstance(item, (list, tuple, dict)):
        text = json.dumps(item, ensure_ascii=False)
    else:
        text = str(item)
    text = " ".join(text.split())
    return text[:max_chars]


def sheet_rows(path: Path, sheet_name: str) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        iterator = ws.iter_rows(values_only=True)
        try:
            headers = [str(item or "").strip() for item in next(iterator)]
        except StopIteration:
            return {}
        result: dict[str, dict[str, Any]] = {}
        for cells in iterator:
            raw = {headers[index]: cells[index] if index < len(cells) else "" for index in range(len(headers))}
            md5 = clean_md5(raw.get("md5"))
            if md5:
                raw["md5"] = md5
                result[md5] = raw
        return result
    finally:
        wb.close()


def first_existing_sheet_rows(path: Path, sheet_names: list[str]) -> dict[str, dict[str, Any]]:
    for sheet_name in sheet_names:
        rows = sheet_rows(path, sheet_name)
        if rows:
            return rows
    return {}


def convert_record(profile: dict[str, Any], network: dict[str, Any], source_file: str, label: str = "") -> dict[str, Any]:
    merged = {**network, **profile}
    md5 = clean_md5(value(merged, "md5"))
    domains = compact_list_text(value(network, "domains"))
    top_domains = compact_list_text(value(network, "top_domains"))
    ips = compact_list_text(value(network, "ips"))
    control_url = domains or top_domains or ips
    plugins = compact_list_text(value(merged, "plugins"))
    permissions = compact_list_text(value(merged, "permissions"))
    code_fuscator = normalize_bool(value(merged, "codeFuscator"))
    source_malicious = normalize_bool(value(network, "fraud_count"))
    if label == "malicious":
        source_malicious = "true"
    elif label == "white":
        source_malicious = "false"
    return {
        "md5": md5,
        "sample_id": md5,
        "app_name": value(merged, "appName", "input_appName"),
        "package_name": value(merged, "pkgName"),
        "sha1": value(merged, "fileSha1"),
        "sha256": value(merged, "fileSha256"),
        "version_name": value(merged, "versionName", "version"),
        "version_code": value(merged, "versionCode"),
        "file_size": value(merged, "size"),
        "file_type": value(merged, "fileType"),
        "signature_status": normalize_signature_status(merged),
        "certificate_fingerprint": value(merged, "certMd5", "signMd5", "signatureMd5"),
        "cert_sha1": value(merged, "certSha1", "signSha1"),
        "cert_sha256": value(merged, "certSha256", "signSha256"),
        "certificate_owner": value(merged, "certOwner", "certificateOwner", "owner"),
        "certificate_developer": value(merged, "certDeveloper", "developer", "appDeveloper"),
        "permissions": permissions,
        "plugins": plugins,
        "sdk_list": plugins,
        "packer": code_fuscator,
        "code_fuscator": code_fuscator,
        "unshell_info": compact_list_text(value(merged, "unshellInfo")),
        "fake_app": normalize_bool(value(merged, "fakeApp")),
        "genuine": normalize_bool(value(merged, "genuine")),
        "official_app_name": value(merged, "genuineAppName", "officialAppName"),
        "rebuild_type": value(merged, "rebuildType"),
        "icon_md5": value(merged, "iconMd5", "iconFinger"),
        "virus_name": value(merged, "virusName", "fraudName"),
        "risk_type": value(merged, "riskType"),
        "virus_description": compact_list_text(value(merged, "virusDescription", "description")),
        "fraud_flag": normalize_bool(value(network, "fraud_count", "isFraud")),
        "fraud_type_info": compact_list_text(value(merged, "fraudTypeInfo")),
        "fraud_name": value(merged, "fraudName"),
        "fraud_category": value(merged, "fraudCategory"),
        "fraud_category_big": value(merged, "input_fraudGaType"),
        "fraud_category_small": value(merged, "input_fraudGaSubType"),
        "control_url": control_url,
        "download_url": value(merged, "downloadUrl"),
        "lt_urls": compact_list_text(value(merged, "ltUrls")),
        "dynamic_nets": compact_list_text(value(merged, "dynamicNets")),
        "domains": domains,
        "top_domains": top_domains,
        "ips": ips,
        "countries": compact_list_text(value(network, "countries")),
        "operators": compact_list_text(value(network, "operators")),
        "domain_count": value(network, "domain_count"),
        "ip_count": value(network, "ip_count"),
        "network_record_count": value(network, "url_record_count"),
        "source_risk_score": value(merged, "score"),
        "source_malicious": source_malicious,
        "source_violation": source_file,
        "query_time": value(merged, "update_time", "query_time"),
    }


def write_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(OUTPUT_SHEET)
    ws.append(APP_HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in APP_HEADERS])
    wb.save(path)


def has_features(row: dict[str, Any]) -> bool:
    for header in FEATURE_HEADERS:
        item = row.get(header)
        if item not in (None, "", "false", 0):
            return True
    return False


def convert_file(
    source: Path,
    output_dir: Path,
    label: str = "",
    drop_empty_features: bool = False,
) -> tuple[Path, list[dict[str, Any]], int]:
    profiles = first_existing_sheet_rows(source, [PROFILE_SHEET, PROFILE_LITE_SHEET])
    networks = first_existing_sheet_rows(source, [NETWORK_SUMMARY_SHEET, NETWORK_SUMMARY_LITE_SHEET])
    raw_inputs = sheet_rows(source, RAW_INPUT_SHEET)
    rows = []
    dropped = 0
    for md5 in sorted(set(raw_inputs) | set(profiles) | set(networks)):
        row = convert_record(
            {**raw_inputs.get(md5, {}), **profiles.get(md5, {})},
            networks.get(md5, {}),
            source.name,
            label,
        )
        if drop_empty_features and not has_features(row):
            dropped += 1
            continue
        rows.append(row)
    output = output_dir / f"{source.stem}{SINGLE_SUFFIX}"
    write_workbook(output, rows)
    return output, rows, dropped


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert output_app_judgment workbooks to MalApp one-sheet import format.")
    parser.add_argument("--src", help="output_app_judgment directory")
    parser.add_argument("--files", nargs="*", help="explicit xlsx files to convert")
    parser.add_argument("--out", required=True, help="output directory")
    parser.add_argument("--combined-name", default=DEFAULT_COMBINED_NAME)
    parser.add_argument("--label", choices=["", "malicious", "white"], default="")
    parser.add_argument("--limit", type=int, default=0, help="limit combined output rows after de-duplication")
    parser.add_argument("--drop-empty-features", action="store_true", help="drop rows without usable APP judgement features")
    args = parser.parse_args()

    output_dir = Path(args.out)
    output_dir.mkdir(parents=True, exist_ok=True)
    all_rows: list[dict[str, Any]] = []
    summary = []
    if args.files:
        sources = [Path(item) for item in args.files]
    elif args.src:
        sources = sorted(Path(args.src).glob("*.xlsx"))
    else:
        raise SystemExit("Either --src or --files is required.")
    for source in sources:
        output, rows, dropped = convert_file(source, output_dir, args.label, args.drop_empty_features)
        all_rows.extend(rows)
        summary.append({"source": str(source), "output": str(output), "rows": len(rows), "dropped_empty_features": dropped})
        print(f"converted {source.name}: {len(rows)} rows, dropped {dropped} -> {output}", flush=True)
    if args.limit:
        deduped: dict[str, dict[str, Any]] = {}
        for row in all_rows:
            md5 = str(row.get("md5") or "")
            if md5 and md5 not in deduped:
                deduped[md5] = row
            if len(deduped) >= args.limit:
                break
        all_rows = list(deduped.values())
    combined = output_dir / args.combined_name
    write_workbook(combined, all_rows)
    summary.append({"source": "ALL", "output": str(combined), "rows": len(all_rows)})
    (output_dir / SUMMARY_NAME).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"combined: {len(all_rows)} rows -> {combined}", flush=True)


if __name__ == "__main__":
    main()
