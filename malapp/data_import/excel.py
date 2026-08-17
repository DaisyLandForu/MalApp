from __future__ import annotations

from io import BytesIO
from typing import Any

from malapp.application.dashboard import import_feature_records


def open_workbook(content: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("桌面程序未包含 Excel 读取组件") from exc
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取 Excel 文件：{exc}") from exc


def excel_preview(content: bytes, sheet_name: str = "", header_row: int = 1) -> dict[str, Any]:
    workbook = open_workbook(content)
    try:
        selected = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[selected]
        header_row = max(1, int(header_row))
        headers = [
            clean_cell(value) or f"第{index + 1}列"
            for index, value in enumerate(
                next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
            )
        ]
        preview_rows = []
        for values in sheet.iter_rows(
            min_row=header_row + 1,
            max_row=min(sheet.max_row, header_row + 5),
            values_only=True,
        ):
            preview_rows.append([clean_cell(value) for value in values[: min(len(headers), 12)]])
        return {
            "sheets": workbook.sheetnames,
            "selected_sheet": selected,
            "total_rows": max(0, sheet.max_row - header_row),
            "total_columns": sheet.max_column,
            "headers": headers,
            "preview_rows": preview_rows,
        }
    finally:
        workbook.close()


def import_excel(
    content: bytes,
    *,
    sheet_name: str = "",
    header_row: int = 1,
    start_row: int = 2,
    limit: int = 100,
    source: str = "excel_upload",
) -> dict[str, Any]:
    workbook = open_workbook(content)
    try:
        selected = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[selected]
        header_row = max(1, int(header_row))
        start_row = max(header_row + 1, int(start_row))
        limit = max(1, min(int(limit), 100_000))
        header_values = next(
            sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
            (),
        )
        headers = unique_headers(header_values)
        items: list[dict[str, Any]] = []
        skipped_blank = 0
        for values in sheet.iter_rows(min_row=start_row, values_only=True):
            record = {
                headers[index]: clean_cell(value)
                for index, value in enumerate(values[: len(headers)])
                if clean_cell(value) not in ("", None)
            }
            if not record:
                skipped_blank += 1
                continue
            items.append(record)
            if len(items) >= limit:
                break
        if not items:
            raise ValueError("所选范围内没有可导入的数据")
        result = import_feature_records(items, source=source, payload_format="json")
        result.update(
            {
                "sheet": selected,
                "header_row": header_row,
                "start_row": start_row,
                "requested_limit": limit,
                "skipped_blank_rows": skipped_blank,
            }
        )
        return result
    finally:
        workbook.close()


def unique_headers(values: tuple[Any, ...]) -> list[str]:
    result = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        base = str(clean_cell(value) or f"第{index + 1}列").strip()
        seen[base] = seen.get(base, 0) + 1
        result.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return result


def clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except (TypeError, ValueError):
            pass
    return value
