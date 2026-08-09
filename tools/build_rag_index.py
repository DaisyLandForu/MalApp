from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from engine.pipeline import DB_PATH  # noqa: E402
from engine.rag import RAG_DB_PATH, add_document, init_rag_db, rag_status  # noqa: E402


def doc_id(prefix: str, *parts: Any) -> str:
    raw = "|".join(str(part) for part in parts)
    digest = hashlib.sha1(raw.encode("utf-8", "ignore")).hexdigest()[:20]
    return f"{prefix}:{digest}"


def ingest_reports(limit: int | None = None) -> int:
    if not DB_PATH.exists():
        return 0
    count = 0
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        sql = "SELECT id, sample_id, verdict, final_score, risk_level, payload_json FROM judgements ORDER BY created_at DESC"
        if limit:
            sql += f" LIMIT {int(limit)}"
        for row in conn.execute(sql):
            try:
                report = json.loads(row["payload_json"])
            except Exception:
                continue
            sample = report.get("sample", {})
            decision = report.get("decision", {})
            blocks = report.get("evidence_blocks", [])
            evidence_text = []
            for block in blocks[:4] if isinstance(blocks, list) else []:
                if not isinstance(block, dict):
                    continue
                claim = safe_text(block.get("claim"))
                evidence_items = [safe_text(x) for x in block.get("evidence", [])[:3]]
                evidence_items = [x for x in evidence_items if x]
                if not claim and not evidence_items:
                    continue
                evidence_text.append(
                    f"agent={block.get('agent')}; claim={claim}; evidence={'; '.join(evidence_items)}"
                )
            title = (
                f"historical judgement case {safe_text(sample.get('app_name')) or row['sample_id']} "
                f"{decision.get('verdict') or row['verdict']}"
            )
            content = "\n".join(
                [
                    f"sample={sample.get('md5') or row['sample_id']}; app={safe_text(sample.get('app_name'))}; package={safe_text(sample.get('package_name'))}",
                    f"final_verdict={decision.get('verdict') or row['verdict']}; risk={decision.get('risk_level') or row['risk_level']}; score={decision.get('final_score') or row['final_score']}",
                    "evidence_blocks:\n" + "\n".join(evidence_text),
                ]
            )
            add_document(
                doc_id=doc_id("case", row["id"], row["sample_id"]),
                source_type="historical_case",
                source_name="judgements",
                title=title,
                content=content,
                metadata={"report_id": row["id"], "md5": sample.get("md5"), "label": decision.get("verdict")},
            )
            count += 1
    return count


def ingest_manual_labels(limit: int | None = None) -> int:
    if not DB_PATH.exists():
        return 0
    count = 0
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        sql = "SELECT md5, label, source_file, conflict_type, raw_json FROM manual_labels ORDER BY imported_at DESC"
        if limit:
            sql += f" LIMIT {int(limit)}"
        for row in conn.execute(sql):
            raw = _loads(row["raw_json"], {})
            title = f"manual conflict case {row['label']} {row['md5']}"
            content = "\n".join(
                [
                    f"md5={row['md5']}; manual_label={row['label']}; conflict_type={row['conflict_type']}",
                    f"source_file={row['source_file']}",
                    "raw_fields_summary=" + json.dumps(raw, ensure_ascii=False, sort_keys=True)[:1800],
                ]
            )
            add_document(
                doc_id=doc_id("manual", row["md5"], row["label"], row["source_file"]),
                source_type="manual_case",
                source_name=Path(str(row["source_file"])).name,
                title=title,
                content=content,
                metadata={"md5": row["md5"], "label": row["label"], "conflict_type": row["conflict_type"]},
            )
            count += 1
    return count


def ingest_misp_galaxy(path: Path, limit: int | None = None) -> int:
    if not path.exists():
        return 0
    count = 0
    for file in sorted(path.rglob("*.json")):
        if limit and count >= limit:
            break
        try:
            data = json.loads(file.read_text(encoding="utf-8"))
        except Exception:
            continue
        values = data.get("values") if isinstance(data, dict) else None
        if not isinstance(values, list):
            continue
        galaxy_name = str(data.get("name") or file.stem)
        for item in values:
            if limit and count >= limit:
                break
            if not isinstance(item, dict):
                continue
            meta = item.get("meta") if isinstance(item.get("meta"), dict) else {}
            synonyms = meta.get("synonyms") if isinstance(meta.get("synonyms"), list) else []
            value = item.get("value") or item.get("uuid") or (synonyms[0] if synonyms else "")
            description = item.get("description") or meta.get("description") or ""
            title = f"MISP Galaxy {galaxy_name}: {value}"
            content = "\n".join(
                [
                    f"knowledge_base={galaxy_name}",
                    f"name={value}",
                    f"description={description}",
                    "meta=" + json.dumps(meta, ensure_ascii=False, sort_keys=True)[:1400],
                ]
            )
            add_document(
                doc_id=doc_id("misp", file.relative_to(path), value),
                source_type="threat_family_ioc",
                source_name=f"MISP Galaxy/{galaxy_name}",
                title=title,
                content=content,
                metadata={"path": str(file), "galaxy": galaxy_name, "value": value},
            )
            count += 1
    return count


def ingest_docs(path: Path, limit: int | None = None) -> int:
    if not path.exists():
        return 0
    count = 0
    for file in sorted(path.rglob("*.md")):
        if limit and count >= limit:
            break
        if "?" in file.name or "\ufffd" in file.name or file.name.startswith("RAG"):
            continue
        try:
            text = file.read_text(encoding="utf-8")
        except Exception:
            continue
        for index, chunk in enumerate(chunk_text(text)):
            if limit and count >= limit:
                break
            title = f"judgement spec {file.stem} #{index + 1}"
            add_document(
                doc_id=doc_id("doc", file, index, chunk[:80]),
                source_type="judgement_spec",
                source_name=str(file.relative_to(ROOT)),
                title=title,
                content=chunk,
                metadata={"path": str(file), "chunk": index},
            )
            count += 1
    return count


def ingest_app_rag(path: Path, limit: int | None = None) -> int:
    """Ingest the user's APP-RAG folder into the local retrieval database.

    The folder is expected to contain internal APP judgement material in mixed
    formats. We keep the parser dependency-light so the packaged desktop app can
    rebuild the index without downloading extra wheels.
    """
    if not path.exists():
        return 0
    supported = {".docx", ".pptx", ".xlsx", ".pdf", ".txt", ".md", ".csv"}
    count = 0
    for file in sorted(path.rglob("*")):
        if limit and count >= limit:
            break
        if not file.is_file() or file.suffix.lower() not in supported:
            continue
        text = extract_file_text(file)
        if not text:
            continue
        source_type = classify_app_rag_source(file)
        rel_name = safe_relative(file, ROOT)
        for index, chunk in enumerate(chunk_text(text, size=1100, overlap=160)):
            if limit and count >= limit:
                break
            add_document(
                doc_id=doc_id("app_rag", rel_name, index, chunk[:120]),
                source_type=source_type,
                source_name=rel_name,
                title=f"{file.stem} #{index + 1}",
                content=chunk,
                metadata={
                    "path": str(file),
                    "relative_path": rel_name,
                    "extension": file.suffix.lower(),
                    "chunk": index,
                },
            )
            count += 1
    return count


def extract_file_text(file: Path) -> str:
    suffix = file.suffix.lower()
    try:
        if suffix == ".docx":
            return extract_docx_text(file)
        if suffix == ".pptx":
            return extract_pptx_text(file)
        if suffix == ".xlsx":
            return extract_xlsx_text(file)
        if suffix == ".pdf":
            return extract_pdf_text(file)
        if suffix in {".txt", ".md", ".csv"}:
            return file.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""
    return ""


def extract_docx_text(file: Path) -> str:
    pieces: list[str] = []
    with zipfile.ZipFile(file) as archive:
        names = [
            name
            for name in archive.namelist()
            if name.startswith("word/") and name.endswith(".xml")
        ]
        for name in names:
            pieces.extend(xml_text_nodes(archive.read(name)))
    return "\n".join(pieces)


def extract_pptx_text(file: Path) -> str:
    pieces: list[str] = []
    with zipfile.ZipFile(file) as archive:
        names = sorted(
            name
            for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        for name in names:
            slide_text = " ".join(xml_text_nodes(archive.read(name)))
            if slide_text:
                pieces.append(slide_text)
    return "\n".join(pieces)


def extract_xlsx_text(file: Path) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""
    workbook = load_workbook(file, read_only=True, data_only=True)
    pieces: list[str] = []
    for sheet in workbook.worksheets:
        rows = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                rows.append(f"row{row_index}: " + " | ".join(values))
            if len(rows) >= 300:
                rows.append("... sheet truncated at 300 non-empty rows for RAG indexing")
                break
        if rows:
            pieces.append(f"sheet={sheet.title}\n" + "\n".join(rows))
    return "\n\n".join(pieces)


def extract_pdf_text(file: Path) -> str:
    # Prefer installed Python PDF readers if present.
    for module_name in ("pypdf", "PyPDF2"):
        try:
            module = __import__(module_name)
            reader_cls = getattr(module, "PdfReader")
            reader = reader_cls(str(file))
            pages = []
            for page in reader.pages[:80]:
                pages.append(page.extract_text() or "")
            text = "\n".join(pages).strip()
            if text:
                return text
        except Exception:
            pass
    # Then try the common poppler CLI if available on the machine.
    pdftotext = shutil.which("pdftotext")
    if pdftotext:
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "out.txt"
            try:
                subprocess.run(
                    [pdftotext, "-enc", "UTF-8", "-layout", str(file), str(output)],
                    check=False,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    timeout=45,
                )
                if output.exists():
                    text = output.read_text(encoding="utf-8", errors="ignore").strip()
                    if text:
                        return text
            except Exception:
                pass
    # Last resort: extract readable byte runs. It is incomplete but still useful
    # for PDFs that contain uncompressed Chinese/English metadata.
    try:
        raw = file.read_bytes()
        text = raw.decode("utf-8", errors="ignore")
        readable = re.findall(r"[\u4e00-\u9fffA-Za-z0-9，。；：、（）()/_ .:\-]{8,}", text)
        return "\n".join(readable[:500])
    except Exception:
        return ""


def xml_text_nodes(data: bytes) -> list[str]:
    try:
        root = ElementTree.fromstring(data)
    except Exception:
        return []
    pieces: list[str] = []
    for node in root.iter():
        if node.text and node.text.strip():
            pieces.append(" ".join(node.text.split()))
    return pieces


def classify_app_rag_source(file: Path) -> str:
    text = str(file).lower()
    name = file.name.lower()
    if any(key in text for key in ("黑灰产", "黑产", "情报", "ioc", "诈骗", "涉诈", "反诈")):
        return "threat_family_ioc"
    if any(key in text for key in ("正版", "仿冒", "app全景态势", "资产", "功能-名词")):
        return "official_app_asset"
    if any(key in text for key in ("判定", "依据", "标准", "方法", "手册", "培训", "规范")):
        return "judgement_spec"
    if any(key in name for key in ("sdk", "字典", "类别", "标签")):
        return "judgement_spec"
    return "app_rag_document"

def safe_relative(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except Exception:
        return str(path)


def ingest_genuine_sql(path: Path, limit: int | None = None) -> int:
    if not path.exists():
        return 0
    count = 0
    # Conservative parser: extract readable INSERT row fragments without executing SQL.
    pattern = re.compile(r"\(([^()]{20,3000})\)")
    with path.open("r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            if limit and count >= limit:
                break
            if "INSERT" not in line.upper():
                continue
            for match in pattern.finditer(line):
                if limit and count >= limit:
                    break
                row = match.group(1)
                if not any(key in row.lower() for key in ("com.", "android", "app", "md5")):
                    continue
                title = f"official app asset {count + 1}"
                content = "official_app_asset_record=" + row[:2200]
                add_document(
                    doc_id=doc_id("genuine", path.name, count, row[:120]),
                    source_type="official_app_asset",
                    source_name=path.name,
                    title=title,
                    content=content,
                    metadata={"path": str(path), "row_index": count},
                )
                count += 1
    return count


def chunk_text(text: str, size: int = 900, overlap: int = 120) -> list[str]:
    clean = "\n".join(line.strip() for line in text.splitlines() if line.strip())
    if not clean:
        return []
    chunks = []
    start = 0
    while start < len(clean):
        end = min(len(clean), start + size)
        chunks.append(clean[start:end])
        if end == len(clean):
            break
        start = max(end - overlap, start + 1)
    return chunks


def _loads(value: str, fallback: Any) -> Any:
    try:
        return json.loads(value)
    except Exception:
        return fallback


def safe_text(value: Any, limit: int = 600) -> str:
    text = str(value or "")
    bad_markers = ("�", "锛", "鐨", "妯", "鎭", "绐", "涓", "鍙", "旂", "佹")
    if any(marker in text for marker in bad_markers):
        return ""
    return " ".join(text.split())[:limit]


def reset_index() -> None:
    init_rag_db()
    with sqlite3.connect(RAG_DB_PATH) as conn:
        conn.execute("DELETE FROM rag_documents")
        conn.commit()


def main() -> None:
    parser = argparse.ArgumentParser(description="Build local RAG index for MalApp judgement.")
    parser.add_argument("--misp-galaxy-dir", default="")
    parser.add_argument("--genuine-sql", default=str(ROOT / "genuine_new.sql"))
    parser.add_argument("--docs-dir", default=str(ROOT / "docs"))
    parser.add_argument("--app-rag-dir", default=str(ROOT / "APP-RAG"))
    parser.add_argument("--limit-per-source", type=int, default=0)
    parser.add_argument("--skip-reports", action="store_true")
    parser.add_argument("--skip-manual", action="store_true")
    parser.add_argument("--skip-docs", action="store_true")
    parser.add_argument("--skip-app-rag", action="store_true")
    parser.add_argument("--skip-genuine", action="store_true")
    parser.add_argument("--reset", action="store_true")
    args = parser.parse_args()

    init_rag_db()
    if args.reset:
        reset_index()
    limit = args.limit_per_source or None
    counts = {}
    if not args.skip_reports:
        counts["historical_case"] = ingest_reports(limit)
    if not args.skip_manual:
        counts["manual_case"] = ingest_manual_labels(limit)
    if args.misp_galaxy_dir:
        counts["threat_family_ioc"] = ingest_misp_galaxy(Path(args.misp_galaxy_dir), limit)
    if not args.skip_docs:
        counts["judgement_spec"] = ingest_docs(Path(args.docs_dir), limit)
    if not args.skip_app_rag:
        counts["app_rag"] = ingest_app_rag(Path(args.app_rag_dir), limit)
    if not args.skip_genuine:
        counts["official_app_asset"] = ingest_genuine_sql(Path(args.genuine_sql), limit)
    print(json.dumps({"counts": counts, "status": rag_status(), "database": str(RAG_DB_PATH)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
