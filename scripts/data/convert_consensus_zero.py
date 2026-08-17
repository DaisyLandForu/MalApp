from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from scripts.data.convert_manual_conflicts import clean_md5, convert_record, row_dict, write_workbook


def is_zero_score(value: Any) -> bool:
    try:
        return float(str(value).strip()) == 0.0
    except Exception:
        return False


def load_360_zero_rows(path: Path) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        iterator = ws.iter_rows(values_only=True)
        headers = [str(item or "").strip() for item in next(iterator)]
        for row in iterator:
            item = row_dict(headers, row)
            md5 = clean_md5(item.get("MD5"))
            if md5 and is_zero_score(item.get("score")):
                rows[md5] = item
    finally:
        wb.close()
    return rows


def collect_consensus_zero(engine360: Path, enginecm: Path, limit: int) -> list[dict[str, Any]]:
    rows360 = load_360_zero_rows(engine360)
    output: list[dict[str, Any]] = []
    seen: set[str] = set()
    wb = load_workbook(enginecm, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        iterator = ws.iter_rows(values_only=True)
        headers = [str(item or "").strip() for item in next(iterator)]
        for row in iterator:
            itemcm = row_dict(headers, row)
            md5 = clean_md5(itemcm.get("MD5"))
            if not md5 or md5 in seen or md5 not in rows360:
                continue
            if not is_zero_score(itemcm.get("score")):
                continue
            item360 = rows360[md5]
            manual = {
                "md5": md5,
                "manual_review_result": "0",
                "conflict_type": "360_CM_same_score_0",
                "engine_360_type": str(item360.get("\u7c7b\u578b") or ""),
                "engine_360_score": str(item360.get("score") or "0"),
                "engine_360_virus_name": str(item360.get("\u75c5\u6bd2\u540d\u79f0") or ""),
                "engine_cm_type": str(itemcm.get("\u7c7b\u578b") or ""),
                "engine_cm_score": str(itemcm.get("score") or "0"),
                "engine_cm_virus_name": str(itemcm.get("\u75c5\u6bd2\u540d\u79f0") or ""),
                "app_name_360": str(item360.get("\u5e94\u7528\u540d\u79f0") or ""),
                "app_name_cm": str(itemcm.get("\u5e94\u7528\u540d\u79f0") or ""),
                "label_source": "360_cm_same_score_0",
            }
            output.append(convert_record(manual, item360, itemcm))
            seen.add(md5)
            if len(output) >= limit:
                break
    finally:
        wb.close()
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--engine-360", required=True)
    parser.add_argument("--engine-cm", required=True)
    parser.add_argument("--limit", type=int, default=1200)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    rows = collect_consensus_zero(Path(args.engine_360), Path(args.engine_cm), args.limit)
    output = Path(args.out)
    output.parent.mkdir(parents=True, exist_ok=True)
    write_workbook(output, rows)
    summary = {
        "rule": "360 score == 0 and cm score == 0",
        "requested_limit": args.limit,
        "output_rows": len(rows),
        "output": str(output),
    }
    summary_path = output.with_suffix(".summary.json")
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
